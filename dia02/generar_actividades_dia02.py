#!/usr/bin/env python3
"""Genera los 3 archivos Excel de actividades para Día 2 — INDICE+COINCIDIR y Funciones Condicionales."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

BASE = os.path.dirname(os.path.abspath(__file__))

PURPLE   = '7C6FF0'
GREEN    = '1AB87A'
AMBER    = 'F0B429'
RED      = 'E05050'
BLUE     = '6FA8DC'
WHITE    = 'FFFFFF'
DARK     = '0D1B2A'
YELLOW_FILL = 'FFF9C4'
PURPLE_LIGHT = 'F0EEFF'
GREEN_LIGHT  = 'E8F8F0'

def _side(): return Side(style='thin', color='CCCCCC')
def _border(): s = _side(); return Border(left=s, right=s, top=s, bottom=s)

def _fill(color): return PatternFill('solid', fgColor=color)

def _hdr(ws, row, col_start, col_end, bg=PURPLE, fg=WHITE, size=10, bold=True):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=bold, color=fg, size=size)
        cell.fill = _fill(bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = _border()

def _cell(ws, row, col, value=None, bold=False, color='333333', size=10,
          fill=None, align='left', wrap=False, fmt=None, italic=False):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font = Font(bold=bold, color=color, size=size, italic=italic)
    if fill:
        cell.fill = _fill(fill)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    cell.border = _border()
    if fmt:
        cell.number_format = fmt
    return cell

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _freeze(ws, cell='A2'):
    ws.freeze_panes = cell

def _title_row(ws, row, text, cols, color=PURPLE):
    ws.cell(row=row, column=1).value = text
    ws.cell(row=row, column=1).font = Font(bold=True, size=13, color=color)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.row_dimensions[row].height = 22

# ─────────────────────────────────────────────────────────────────────────────
# ARCHIVO 1: ACTIVIDAD INDICE + COINCIDIR
# ─────────────────────────────────────────────────────────────────────────────
def actividad_indice_coincidir():
    wb = openpyxl.Workbook()

    # ── HOJA 1: GUIA ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'GUIA'
    ws.column_dimensions['A'].width = 90
    ws.row_dimensions[1].height = 35

    guia_rows = [
        (1,  'ACTIVIDAD 1 — INDICE + COINCIDIR',                           True, 16, PURPLE, None),
        (2,  'DÍA 2 | Módulo 11 — Ofimática Avanzada | INTEBIEN S.A.S.',   False, 10, '888888', None),
        (3,  '', False, 10, '333333', None),
        (4,  '¿QUÉ APRENDERÁS?', True, 12, PURPLE, None),
        (5,  'Aprenderás a combinar INDICE + COINCIDIR para hacer búsquedas flexibles en cualquier dirección.', False, 10, '333333', None),
        (6,  'Esta combinación es el estándar profesional que supera a BUSCARV en todos los escenarios avanzados.', False, 10, '333333', None),
        (7,  '', False, 10, '333333', None),
        (8,  'ESTRUCTURA DE ESTA ACTIVIDAD', True, 12, PURPLE, None),
        (9,  '  Hoja 2 — Base_Empleados   : Los datos que usarás como fuente de búsqueda (20 empleados)', False, 10, '333333', None),
        (10, '  Hoja 3 — PASO_1_COINCIDIR : Practica COINCIDIR sola (encontrar posición de un valor)', False, 10, '333333', None),
        (11, '  Hoja 4 — PASO_2_INDICE    : Practica INDICE sola (obtener un valor por posición)', False, 10, '333333', None),
        (12, '  Hoja 5 — PASO_3_COMBINADAS: Combina ambas — el poder de INDICE+COINCIDIR', False, 10, '333333', None),
        (13, '  Hoja 6 — RETO_Bidireccional: Búsqueda en filas Y columnas (nivel avanzado)', False, 10, '333333', None),
        (14, '', False, 10, '333333', None),
        (15, 'INSTRUCCIONES GENERALES', True, 12, PURPLE, None),
        (16, '  1. Lee cada hoja en orden (PASO_1 → PASO_2 → PASO_3 → RETO)', False, 10, '333333', None),
        (17, '  2. Las celdas en AMARILLO son donde debes escribir tus fórmulas', False, 10, '333333', None),
        (18, '  3. Las celdas en VERDE claro son el resultado esperado (para verificar)', False, 10, '333333', None),
        (19, '  4. Empieza con los ejemplos resueltos, luego intenta los ejercicios vacíos', False, 10, '333333', None),
        (20, '  5. Si te trabas, revisa la fórmula del ejemplo de arriba como pista', False, 10, '333333', None),
        (21, '', False, 10, '333333', None),
        (22, 'FÓRMULAS CLAVE', True, 12, PURPLE, None),
        (23, '  COINCIDIR:          =COINCIDIR(valor_buscado, rango_busqueda, 0)', False, 10, '0070C0', None),
        (24, '  INDICE:             =INDICE(rango_resultado, numero_fila)', False, 10, '0070C0', None),
        (25, '  INDICE+COINCIDIR:   =INDICE(col_resultado, COINCIDIR(valor, col_busqueda, 0))', False, 10, '0070C0', None),
        (26, '  Bidireccional:      =INDICE(tabla, COINCIDIR(fila,col_filas,0), COINCIDIR(col,fila_cols,0))', False, 10, '0070C0', None),
        (27, '', False, 10, '333333', None),
        (28, 'ENTREGA: Guarda como  Actividad1_IC_TuNombre.xlsx', True, 11, RED, None),
    ]
    for row_i, (row, text, bold, size, color, fill) in enumerate(guia_rows):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=bold, size=size, color=color)
        if fill:
            c.fill = _fill(fill)

    # ── HOJA 2: BASE_EMPLEADOS ─────────────────────────────────────────────────
    ws2 = wb.create_sheet('Base_Empleados')
    _title_row(ws2, 1, 'BASE DE DATOS — EMPLEADOS ADMIN CARIBE S.A.S. (Fuente de búsqueda)', 8)
    ws2.merge_cells('A1:H1')

    headers = ['ID', 'NOMBRE_COMPLETO', 'DEPARTAMENTO', 'CARGO', 'CIUDAD', 'SALARIO', 'ANTIGUEDAD', 'BONO']
    for i, h in enumerate(headers, 1):
        ws2.cell(row=2, column=i).value = h
    _hdr(ws2, 2, 1, 8)

    empleados = [
        ('E001', 'Ana Torres Herrera',      'Administración', 'Auxiliar Admin',       'Barranquilla', 1850000, 2, 0),
        ('E002', 'Carlos Martínez López',   'Ventas',         'Vendedor Senior',       'Barranquilla', 2900000, 5, 290000),
        ('E003', 'María Rodríguez Díaz',    'Contabilidad',   'Auxiliar Contable',     'Barranquilla', 1950000, 3, 0),
        ('E004', 'Juan García Peña',        'Logística',      'Coordinador Logístico', 'Bogotá',       3300000, 7, 330000),
        ('E005', 'Laura Díaz Castro',       'Gerencia',       'Asistente de Gerencia', 'Barranquilla', 2600000, 4, 260000),
        ('E006', 'Pedro Herrera Ruiz',      'Ventas',         'Vendedor Junior',       'Cartagena',    1750000, 1, 0),
        ('E007', 'Sandra Gómez Vargas',     'Ventas',         'Vendedora Senior',      'Bogotá',       3000000, 6, 300000),
        ('E008', 'Miguel López Silva',      'Logística',      'Auxiliar Logístico',    'Barranquilla', 1700000, 2, 0),
        ('E009', 'Diana Moreno Castillo',   'RRHH',           'Analista RRHH',         'Barranquilla', 2200000, 3, 0),
        ('E010', 'Luis Vargas Ospina',      'Contabilidad',   'Contador Jr',           'Medellín',     2400000, 3, 0),
        ('E011', 'Roberto Silva Mejía',     'IT',             'Soporte Técnico',       'Bogotá',       2500000, 4, 0),
        ('E012', 'Valeria Gómez Arbeláez',  'Gerencia',       'Gerente Comercial',     'Barranquilla', 5800000, 9, 580000),
        ('E013', 'Andrés Castro Blanco',    'Ventas',         'Vendedor Senior',       'Barranquilla', 2800000, 5, 280000),
        ('E014', 'Patricia Ruiz Molina',    'Administración', 'Recepcionista',         'Barranquilla', 1600000, 1, 0),
        ('E015', 'Felipe Torres Jiménez',   'Logística',      'Jefe de Bodega',        'Bogotá',       3100000, 6, 310000),
        ('E016', 'Carolina Pérez Ávila',    'Contabilidad',   'Contadora Senior',      'Barranquilla', 3500000, 8, 350000),
        ('E017', 'Hernando Morales Cano',   'Ventas',         'Vendedor Junior',       'Cartagena',    1700000, 1, 0),
        ('E018', 'Lucía Fernández Suárez',  'RRHH',           'Jefe RRHH',             'Barranquilla', 4200000, 7, 420000),
        ('E019', 'Camilo Ríos Becerra',     'IT',             'Desarrollador',         'Bogotá',       4500000, 5, 450000),
        ('E020', 'Natalia Vega Quintero',   'Gerencia',       'Gerente General',       'Barranquilla', 8500000, 12, 850000),
    ]

    alts = [PURPLE_LIGHT, WHITE]
    for r, row in enumerate(empleados, 3):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c == 6:
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif c == 8:
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = _fill(alts[r % 2])
            cell.border = _border()

    ws2.cell(row=24, column=1).value = '← No modifiques esta hoja. Es solo la fuente de datos.'
    ws2.cell(row=24, column=1).font = Font(italic=True, color='888888', size=9)
    ws2.merge_cells('A24:H24')

    _widths(ws2, [6, 26, 16, 24, 14, 13, 13, 12])
    _freeze(ws2)
    ws2.auto_filter.ref = 'A2:H2'

    # ── HOJA 3: PASO_1_COINCIDIR ──────────────────────────────────────────────
    ws3 = wb.create_sheet('PASO_1_COINCIDIR')
    _title_row(ws3, 1, 'PASO 1 — Función COINCIDIR: Encuentra la posición de un valor en un rango', 5)
    ws3.merge_cells('A1:E1')

    # Concepto
    ws3.cell(row=2, column=1).value = '¿QUÉ HACE COINCIDIR?  →  Devuelve el NÚMERO DE FILA donde se encuentra un valor dentro de un rango de una sola columna.'
    ws3.cell(row=2, column=1).font = Font(bold=True, size=10, color='333333')
    ws3.merge_cells('A2:E2')

    ws3.cell(row=3, column=1).value = 'SINTAXIS:  =COINCIDIR( valor_buscado , rango_busqueda , 0 )   ← El 0 significa "coincidencia exacta" — SIEMPRE usar 0.'
    ws3.cell(row=3, column=1).font = Font(size=10, color='0070C0', bold=False)
    ws3.merge_cells('A3:E3')
    ws3.row_dimensions[3].height = 16

    # Tabla de ejercicios
    headers3 = ['EMPLEADO A BUSCAR', 'RANGO DE BÚSQUEDA', 'TU FÓRMULA COINCIDIR', 'POSICIÓN (respuesta)', 'NOTA EXPLICATIVA']
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=5, column=i).value = h
    _hdr(ws3, 5, 1, 5)

    ejercicios = [
        # (nombre, col_busqueda, posicion_esperada, nota)
        ('Ana Torres Herrera',     'Base_Empleados!B3:B22', 1,  'Ejemplo resuelto — Ana es la primera empleada'),
        ('Carlos Martínez López',  'Base_Empleados!B3:B22', 2,  'Ejemplo resuelto — Carlos es el segundo'),
        ('Sandra Gómez Vargas',    'Base_Empleados!B3:B22', 7,  'Tu turno — escribe la fórmula en amarillo'),
        ('Valeria Gómez Arbeláez', 'Base_Empleados!B3:B22', 12, 'Tu turno'),
        ('Felipe Torres Jiménez',  'Base_Empleados!B3:B22', 15, 'Tu turno'),
        ('Natalia Vega Quintero',  'Base_Empleados!B3:B22', 20, 'Tu turno — ¿qué número esperas?'),
    ]

    for r, (nombre, rango, pos, nota) in enumerate(ejercicios, 6):
        _cell(ws3, r, 1, nombre, fill=None)
        _cell(ws3, r, 2, f'=Base_Empleados!B3:B22', color='888888', italic=True)
        ws3.cell(row=r, column=2).value = 'Base_Empleados!B3:B22'
        ws3.cell(row=r, column=2).font = Font(size=9, color='888888', italic=True)
        ws3.cell(row=r, column=2).border = _border()

        if r <= 7:  # Primeros 2: resueltos como ejemplo
            formula = f'=COINCIDIR("{nombre}",Base_Empleados!$B$3:$B$22,0)'
            _cell(ws3, r, 3, formula, color='0070C0', fill=GREEN_LIGHT)
        else:  # El resto: celda amarilla vacía para estudiante
            _cell(ws3, r, 3, '← Escribe aquí tu fórmula', color='999999', fill=YELLOW_FILL, italic=True)

        _cell(ws3, r, 4, pos, align='center', fill=GREEN_LIGHT, color=GREEN, bold=True)
        _cell(ws3, r, 5, nota, color='555555')

    ws3.cell(row=13, column=1).value = ''
    ws3.cell(row=14, column=1).value = 'TAREA ADICIONAL: Busca el ID del empleado que tiene salario = $8,500,000'
    ws3.cell(row=14, column=1).font = Font(bold=True, color=AMBER, size=10)
    ws3.cell(row=14, column=5).value = 'Pista: =COINCIDIR(8500000,Base_Empleados!$F$3:$F$22,0)'
    ws3.cell(row=14, column=5).font = Font(color='888888', size=9, italic=True)
    ws3.merge_cells('A14:D14')

    _widths(ws3, [28, 28, 45, 20, 40])

    # ── HOJA 4: PASO_2_INDICE ─────────────────────────────────────────────────
    ws4 = wb.create_sheet('PASO_2_INDICE')
    _title_row(ws4, 1, 'PASO 2 — Función INDICE: Devuelve el valor en una posición específica', 5)
    ws4.merge_cells('A1:E1')

    ws4.cell(row=2, column=1).value = '¿QUÉ HACE INDICE?  →  Devuelve el VALOR que está en la fila N de un rango. Le das la posición, él te da el contenido.'
    ws4.cell(row=2, column=1).font = Font(bold=True, size=10, color='333333')
    ws4.merge_cells('A2:E2')

    ws4.cell(row=3, column=1).value = 'SINTAXIS:  =INDICE( rango_resultado , numero_fila )   ← el número de fila es lo que encontró COINCIDIR en el paso anterior.'
    ws4.cell(row=3, column=1).font = Font(size=10, color='0070C0')
    ws4.merge_cells('A3:E3')

    headers4 = ['NÚMERO DE FILA (dato)', 'COLUMNA A CONSULTAR', 'TU FÓRMULA INDICE', 'RESULTADO ESPERADO', 'NOTA']
    for i, h in enumerate(headers4, 1):
        ws4.cell(row=5, column=i).value = h
    _hdr(ws4, 5, 1, 5)

    ejercicios4 = [
        (1,  'Base_Empleados!B3:B22', 'Ana Torres Herrera',      'Fila 1 → primer empleado de la lista'),
        (4,  'Base_Empleados!B3:B22', 'Juan García Peña',         'Fila 4 → cuarto empleado'),
        (7,  'Base_Empleados!F3:F22', '$3,000,000',               'Fila 7 → salario del séptimo empleado'),
        (12, 'Base_Empleados!C3:C22', 'Gerencia',                 'Tu turno — ¿qué departamento es el 12?'),
        (20, 'Base_Empleados!E3:E22', 'Barranquilla',             'Tu turno — ¿en qué ciudad está el 20?'),
        (15, 'Base_Empleados!F3:F22', '$3,100,000',               'Tu turno — ¿cuánto gana el empleado 15?'),
    ]

    for r, (fila, rango, resultado, nota) in enumerate(ejercicios4, 6):
        _cell(ws4, r, 1, fila, align='center', bold=True, color=PURPLE, fill=PURPLE_LIGHT)
        ws4.cell(row=r, column=2).value = rango.replace('Base_Empleados!', '').replace('$', '')
        ws4.cell(row=r, column=2).font = Font(size=9, color='888888', italic=True)
        ws4.cell(row=r, column=2).border = _border()

        if r <= 8:  # resueltos
            formula_rng = rango.replace('!', '!')
            formula = f'=INDICE({rango},{fila})'
            _cell(ws4, r, 3, formula, color='0070C0', fill=GREEN_LIGHT)
        else:
            _cell(ws4, r, 3, '← Escribe aquí tu fórmula', color='999999', fill=YELLOW_FILL, italic=True)

        _cell(ws4, r, 4, resultado, align='center', fill=GREEN_LIGHT, color=GREEN, bold=True)
        _cell(ws4, r, 5, nota, color='555555')

    _widths(ws4, [16, 30, 42, 22, 40])

    # ── HOJA 5: PASO_3_COMBINADAS ─────────────────────────────────────────────
    ws5 = wb.create_sheet('PASO_3_COMBINADAS')
    _title_row(ws5, 1, 'PASO 3 — INDICE + COINCIDIR juntas: La búsqueda profesional completa', 6)
    ws5.merge_cells('A1:F1')

    ws5.cell(row=2, column=1).value = 'LÓGICA: COINCIDIR encuentra DÓNDE está el valor (posición) → INDICE va a esa posición y trae el dato que necesitas.'
    ws5.cell(row=2, column=1).font = Font(bold=True, size=10, color='333333')
    ws5.merge_cells('A2:F2')

    ws5.cell(row=3, column=1).value = 'FÓRMULA:  =INDICE( columna_resultado , COINCIDIR( valor_buscado , columna_busqueda , 0 ) )'
    ws5.cell(row=3, column=1).font = Font(size=10, color='0070C0', bold=True)
    ws5.merge_cells('A3:F3')

    # Panel de búsqueda
    ws5.cell(row=5, column=1).value = 'PANEL DE BÚSQUEDA — escribe un nombre o ID y las fórmulas traen todo lo demás'
    ws5.cell(row=5, column=1).font = Font(bold=True, size=11, color=PURPLE)
    ws5.merge_cells('A5:F5')

    ws5.cell(row=6, column=1).value = 'Escribe el nombre del empleado:'
    ws5.cell(row=6, column=1).font = Font(bold=True, size=10)
    ws5.cell(row=6, column=2).value = 'Carlos Martínez López'
    ws5.cell(row=6, column=2).fill = _fill(YELLOW_FILL)
    ws5.cell(row=6, column=2).font = Font(bold=True, size=10, color=PURPLE)
    ws5.cell(row=6, column=2).border = _border()
    ws5.merge_cells('B6:D6')
    ws5.cell(row=6, column=5).value = '← Cambia este nombre para ver los resultados'
    ws5.cell(row=6, column=5).font = Font(italic=True, color='888888', size=9)
    ws5.merge_cells('E6:F6')

    # Resultados del panel
    resultados = [
        ('Departamento',  f'=IFERROR(INDICE(Base_Empleados!$C$3:$C$22,COINCIDIR($B$6,Base_Empleados!$B$3:$B$22,0)),"No encontrado")'),
        ('Cargo',         f'=IFERROR(INDICE(Base_Empleados!$D$3:$D$22,COINCIDIR($B$6,Base_Empleados!$B$3:$B$22,0)),"No encontrado")'),
        ('Ciudad',        f'=IFERROR(INDICE(Base_Empleados!$E$3:$E$22,COINCIDIR($B$6,Base_Empleados!$B$3:$B$22,0)),"No encontrado")'),
        ('Salario',       f'=IFERROR(INDICE(Base_Empleados!$F$3:$F$22,COINCIDIR($B$6,Base_Empleados!$B$3:$B$22,0)),0)'),
        ('Antigüedad',    f'=IFERROR(INDICE(Base_Empleados!$G$3:$G$22,COINCIDIR($B$6,Base_Empleados!$B$3:$B$22,0)),"No encontrado")'),
    ]

    for r, (campo, formula) in enumerate(resultados, 8):
        _cell(ws5, r, 1, campo, bold=True, size=10)
        cell = ws5.cell(row=r, column=2, value=formula)
        cell.font = Font(bold=True, size=12, color=PURPLE)
        if 'Salario' in campo:
            cell.number_format = '"$"#,##0'
        cell.fill = _fill(PURPLE_LIGHT)
        cell.border = _border()
        ws5.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

    # Instrucción: buscar por ID (otra dirección)
    ws5.cell(row=15, column=1).value = 'EJERCICIO: Búsqueda por ID'
    ws5.cell(row=15, column=1).font = Font(bold=True, size=11, color=AMBER)
    ws5.merge_cells('A15:F15')

    ws5.cell(row=16, column=1).value = 'Escribe un ID:'
    ws5.cell(row=16, column=1).font = Font(bold=True, size=10)
    ws5.cell(row=16, column=2).value = 'E012'
    ws5.cell(row=16, column=2).fill = _fill(YELLOW_FILL)
    ws5.cell(row=16, column=2).font = Font(bold=True, size=10, color=PURPLE)
    ws5.cell(row=16, column=2).border = _border()

    ws5.cell(row=17, column=1).value = 'Nombre del empleado:'
    _cell(ws5, 17, 2, '← TU FÓRMULA AQUÍ (busca en col A, devuelve col B)', color='999999', fill=YELLOW_FILL, italic=True)
    ws5.merge_cells('B17:D17')

    ws5.cell(row=18, column=1).value = 'Salario:'
    _cell(ws5, 18, 2, '← TU FÓRMULA AQUÍ (busca en col A, devuelve col F)', color='999999', fill=YELLOW_FILL, italic=True)
    ws5.merge_cells('B18:D18')

    ws5.cell(row=20, column=1).value = 'PISTA: =IFERROR(INDICE(Base_Empleados!$B$3:$B$22,COINCIDIR($B$16,Base_Empleados!$A$3:$A$22,0)),"No encontrado")'
    ws5.cell(row=20, column=1).font = Font(size=9, color='888888', italic=True)
    ws5.merge_cells('A20:F20')

    _widths(ws5, [22, 30, 16, 16, 30, 16])

    # ── HOJA 6: RETO_Bidireccional ────────────────────────────────────────────
    ws6 = wb.create_sheet('RETO_Bidireccional')
    _title_row(ws6, 1, 'RETO — Búsqueda Bidireccional: INDICE con dos COINCIDIR', 6)
    ws6.merge_cells('A1:F1')

    ws6.cell(row=2, column=1).value = 'NIVEL AVANZADO: Puedes buscar tanto la FILA como la COLUMNA con COINCIDIR, y INDICE devuelve la intersección.'
    ws6.cell(row=2, column=1).font = Font(bold=True, size=10, color='333333')
    ws6.merge_cells('A2:F2')

    ws6.cell(row=3, column=1).value = 'FÓRMULA: =INDICE(tabla_datos, COINCIDIR(fila_buscada, col_de_filas, 0), COINCIDIR(col_buscada, fila_de_cols, 0))'
    ws6.cell(row=3, column=1).font = Font(size=10, color='0070C0', bold=True)
    ws6.merge_cells('A3:F3')

    # Tabla de ventas por ciudad y mes
    ws6.cell(row=5, column=1).value = 'TABLA DE VENTAS — PRIMER TRIMESTRE 2025'
    ws6.cell(row=5, column=1).font = Font(bold=True, size=11, color=PURPLE)
    ws6.merge_cells('A5:D5')

    ws6.cell(row=6, column=1).value = 'CIUDAD / MES'
    ws6.cell(row=6, column=2).value = 'Enero'
    ws6.cell(row=6, column=3).value = 'Febrero'
    ws6.cell(row=6, column=4).value = 'Marzo'
    _hdr(ws6, 6, 1, 4)

    ventas = [
        ('Barranquilla', 42500000, 45800000, 48200000),
        ('Bogotá',       61000000, 65300000, 68500000),
        ('Cartagena',    24100000, 26800000, 28900000),
        ('Medellín',     31000000, 33500000, 36200000),
    ]
    alts = [PURPLE_LIGHT, WHITE]
    for r, (ciudad, e, f, m) in enumerate(ventas, 7):
        for c, val in enumerate([ciudad, e, f, m], 1):
            cell = ws6.cell(row=r, column=c, value=val)
            if c > 1:
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = _fill(alts[r % 2])
            cell.border = _border()

    # Panel de búsqueda bidireccional
    ws6.cell(row=13, column=1).value = 'PANEL DE BÚSQUEDA BIDIRECCIONAL'
    ws6.cell(row=13, column=1).font = Font(bold=True, size=11, color=AMBER)
    ws6.merge_cells('A13:F13')

    ws6.cell(row=14, column=1).value = 'Ciudad a buscar:'
    ws6.cell(row=14, column=2).value = 'Barranquilla'
    ws6.cell(row=14, column=2).fill = _fill(YELLOW_FILL)
    ws6.cell(row=14, column=2).font = Font(bold=True, color=PURPLE)
    ws6.cell(row=14, column=2).border = _border()
    ws6.cell(row=14, column=3).value = '← Cambia ciudad y mes para ver distintos resultados'
    ws6.cell(row=14, column=3).font = Font(italic=True, color='888888', size=9)
    ws6.merge_cells('C14:F14')

    ws6.cell(row=15, column=1).value = 'Mes a buscar:'
    ws6.cell(row=15, column=2).value = 'Febrero'
    ws6.cell(row=15, column=2).fill = _fill(YELLOW_FILL)
    ws6.cell(row=15, column=2).font = Font(bold=True, color=PURPLE)
    ws6.cell(row=15, column=2).border = _border()

    ws6.cell(row=17, column=1).value = 'Resultado (EJEMPLO ya resuelto):'
    ws6.cell(row=17, column=1).font = Font(bold=True, size=10)
    formula_bidir = '=IFERROR(INDICE($B$7:$D$10,COINCIDIR($B$14,$A$7:$A$10,0),COINCIDIR($B$15,$B$6:$D$6,0)),"Ciudad o mes no encontrado")'
    cell_res = ws6.cell(row=17, column=2, value=formula_bidir)
    cell_res.font = Font(bold=True, size=14, color=PURPLE)
    cell_res.number_format = '"$"#,##0'
    cell_res.fill = _fill(PURPLE_LIGHT)
    cell_res.border = _border()
    ws6.merge_cells('B17:D17')

    ws6.cell(row=19, column=1).value = 'AHORA TÚ: En B21:B23, escribe las fórmulas para responder estas preguntas:'
    ws6.cell(row=19, column=1).font = Font(bold=True, size=10, color=AMBER)
    ws6.merge_cells('A19:F19')

    preguntas = [
        ('¿Cuánto vendió Bogotá en Marzo?',      '=IFERROR(INDICE($B$7:$D$10,COINCIDIR("Bogotá",$A$7:$A$10,0),COINCIDIR("Marzo",$B$6:$D$6,0)),0)'),
        ('¿Cuánto vendió Cartagena en Enero?',   '=IFERROR(INDICE($B$7:$D$10,COINCIDIR("Cartagena",$A$7:$A$10,0),COINCIDIR("Enero",$B$6:$D$6,0)),0)'),
        ('¿Cuánto vendió Medellín en Febrero?',  '=IFERROR(INDICE($B$7:$D$10,COINCIDIR("Medellín",$A$7:$A$10,0),COINCIDIR("Febrero",$B$6:$D$6,0)),0)'),
    ]
    for r, (preg, pista) in enumerate(preguntas, 20):
        _cell(ws6, r, 1, preg, bold=True, size=10)
        _cell(ws6, r, 2, '← TU FÓRMULA AQUÍ', color='999999', fill=YELLOW_FILL, italic=True)
        ws6.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        _cell(ws6, r, 5, f'Pista: {pista[:50]}...', color='AAAAAA', size=8, italic=True)
        ws6.merge_cells(start_row=r, start_column=5, end_row=r, end_column=6)

    _widths(ws6, [22, 22, 16, 16, 40, 16])

    path = os.path.join(BASE, 'actividad_01_INDICE_COINCIDIR.xlsx')
    wb.save(path)
    print(f'  ✓ {os.path.relpath(path)}')


# ─────────────────────────────────────────────────────────────────────────────
# ARCHIVO 2: ACTIVIDAD FUNCIONES CONDICIONALES
# ─────────────────────────────────────────────────────────────────────────────
def actividad_condicionales():
    wb = openpyxl.Workbook()

    # ── HOJA 1: GUIA ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'GUIA'
    ws.column_dimensions['A'].width = 90

    guia_rows = [
        (1,  'ACTIVIDAD 2 — FUNCIONES CONDICIONALES',                              True, 16, GREEN),
        (2,  'DÍA 2 | Módulo 11 — Ofimática Avanzada | INTEBIEN S.A.S.',           False, 10, '888888'),
        (3,  '', False, 10, '333333'),
        (4,  '¿QUÉ APRENDERÁS?', True, 12, GREEN),
        (5,  'Las funciones condicionales analizan grandes bases de datos respondiendo preguntas específicas:', False, 10, '333333'),
        (6,  '  ¿Cuántos empleados son de Barranquilla? → CONTAR.SI', False, 10, '333333'),
        (7,  '  ¿Cuánto suma la nómina del área de Ventas? → SUMAR.SI', False, 10, '333333'),
        (8,  '  ¿Cuál es el salario promedio en Bogotá? → PROMEDIO.SI', False, 10, '333333'),
        (9,  '  ¿Cuánto es la nómina de Ventas en Barranquilla? → SUMAR.SI.CONJUNTO', False, 10, '333333'),
        (10, '', False, 10, '333333'),
        (11, 'ESTRUCTURA DE ESTA ACTIVIDAD', True, 12, GREEN),
        (12, '  Hoja 2 — Datos_Nomina     : Base de datos de 20 empleados (tu fuente)', False, 10, '333333'),
        (13, '  Hoja 3 — SUMAR_SI         : Practica SUMAR.SI paso a paso', False, 10, '333333'),
        (14, '  Hoja 4 — CONTAR_SI        : Practica CONTAR.SI paso a paso', False, 10, '333333'),
        (15, '  Hoja 5 — PROMEDIO_SI      : Practica PROMEDIO.SI paso a paso', False, 10, '333333'),
        (16, '  Hoja 6 — RETO_CONJUNTO    : SUMAR.SI.CONJUNTO con múltiples condiciones', False, 10, '333333'),
        (17, '', False, 10, '333333'),
        (18, 'INSTRUCCIONES GENERALES', True, 12, GREEN),
        (19, '  1. Trabaja cada hoja en orden', False, 10, '333333'),
        (20, '  2. Las celdas en AMARILLO son donde escribes tus fórmulas', False, 10, '333333'),
        (21, '  3. Las celdas en VERDE CLARO muestran el resultado correcto para verificar', False, 10, '333333'),
        (22, '  4. En cada hoja hay 2 ejemplos resueltos y luego ejercicios para ti', False, 10, '333333'),
        (23, '', False, 10, '333333'),
        (24, 'FÓRMULAS CLAVE', True, 12, GREEN),
        (25, '  CONTAR.SI:          =CONTAR.SI( rango_condición , condición )', False, 10, '0070C0'),
        (26, '  SUMAR.SI:           =SUMAR.SI( rango_condición , condición , rango_suma )', False, 10, '0070C0'),
        (27, '  PROMEDIO.SI:        =PROMEDIO.SI( rango_condición , condición , rango_promedio )', False, 10, '0070C0'),
        (28, '  SUMAR.SI.CONJUNTO:  =SUMAR.SI.CONJUNTO( rango_suma , rango_crit1 , crit1 , rango_crit2 , crit2 , ... )', False, 10, '0070C0'),
        (29, '', False, 10, '333333'),
        (30, '  ATENCIÓN — SUMAR.SI vs SUMAR.SI.CONJUNTO:', True, 11, RED),
        (31, '  SUMAR.SI:           el rango de suma va AL FINAL', False, 10, RED),
        (32, '  SUMAR.SI.CONJUNTO:  el rango de suma va AL INICIO', False, 10, RED),
        (33, '', False, 10, '333333'),
        (34, 'ENTREGA: Guarda como  Actividad2_Condicionales_TuNombre.xlsx', True, 11, RED),
    ]
    for row_i, (row, text, bold, size, color) in enumerate(guia_rows):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=bold, size=size, color=color)

    # ── HOJA 2: DATOS_NOMINA ──────────────────────────────────────────────────
    ws2 = wb.create_sheet('Datos_Nomina')
    _title_row(ws2, 1, 'BASE DE DATOS NÓMINA — ADMIN CARIBE S.A.S. (No modificar esta hoja)', 7)
    ws2.merge_cells('A1:G1')

    headers2 = ['ID', 'NOMBRE', 'DEPARTAMENTO', 'CIUDAD', 'TIPO_CARGO', 'SALARIO', 'BONO']
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i).value = h
    _hdr(ws2, 2, 1, 7, bg=GREEN)

    nomina = [
        ('E001', 'Ana Torres',       'Administración', 'Barranquilla', 'Auxiliar',    1850000,  0),
        ('E002', 'Carlos Martínez',  'Ventas',         'Barranquilla', 'Senior',      2900000,  290000),
        ('E003', 'María Rodríguez',  'Contabilidad',   'Barranquilla', 'Auxiliar',    1950000,  0),
        ('E004', 'Juan García',      'Logística',      'Bogotá',       'Profesional', 3300000,  330000),
        ('E005', 'Laura Díaz',       'Gerencia',       'Barranquilla', 'Senior',      2600000,  260000),
        ('E006', 'Pedro Herrera',    'Ventas',         'Cartagena',    'Junior',      1750000,  0),
        ('E007', 'Sandra Gómez',     'Ventas',         'Bogotá',       'Senior',      3000000,  300000),
        ('E008', 'Miguel López',     'Logística',      'Barranquilla', 'Auxiliar',    1700000,  0),
        ('E009', 'Diana Moreno',     'RRHH',           'Barranquilla', 'Profesional', 2200000,  0),
        ('E010', 'Luis Vargas',      'Contabilidad',   'Medellín',     'Junior',      2400000,  0),
        ('E011', 'Roberto Silva',    'IT',             'Bogotá',       'Profesional', 2500000,  0),
        ('E012', 'Valeria Gómez',    'Gerencia',       'Barranquilla', 'Senior',      5800000,  580000),
        ('E013', 'Andrés Castro',    'Ventas',         'Barranquilla', 'Senior',      2800000,  280000),
        ('E014', 'Patricia Ruiz',    'Administración', 'Barranquilla', 'Junior',      1600000,  0),
        ('E015', 'Felipe Torres',    'Logística',      'Bogotá',       'Profesional', 3100000,  310000),
        ('E016', 'Carolina Pérez',   'Contabilidad',   'Barranquilla', 'Senior',      3500000,  350000),
        ('E017', 'Hernando Morales', 'Ventas',         'Cartagena',    'Junior',      1700000,  0),
        ('E018', 'Lucía Fernández',  'RRHH',           'Barranquilla', 'Senior',      4200000,  420000),
        ('E019', 'Camilo Ríos',      'IT',             'Bogotá',       'Senior',      4500000,  450000),
        ('E020', 'Natalia Vega',     'Gerencia',       'Barranquilla', 'Senior',      8500000,  850000),
    ]

    alts = [GREEN_LIGHT, WHITE]
    for r, row in enumerate(nomina, 3):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c in [6, 7]:
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = _fill(alts[r % 2])
            cell.border = _border()

    _widths(ws2, [6, 22, 16, 14, 12, 14, 13])
    _freeze(ws2)
    ws2.auto_filter.ref = 'A2:G2'

    # ── HOJA 3: SUMAR_SI ──────────────────────────────────────────────────────
    ws3 = wb.create_sheet('SUMAR_SI')
    _title_row(ws3, 1, 'SUMAR.SI — Suma de valores que cumplen una condición', 6)
    ws3.merge_cells('A1:F1')

    ws3.cell(row=2, column=1).value = 'SINTAXIS: =SUMAR.SI( rango_donde_está_la_condición , "condición" , rango_de_números_a_sumar )'
    ws3.cell(row=2, column=1).font = Font(bold=True, size=10, color='0070C0')
    ws3.merge_cells('A2:F2')

    ws3.cell(row=3, column=1).value = 'Si el rango_condición y el rango_suma son la misma columna, puedes omitir el tercer argumento.'
    ws3.cell(row=3, column=1).font = Font(size=10, color='555555', italic=True)
    ws3.merge_cells('A3:F3')

    headers3 = ['#', 'PREGUNTA', 'RANGO CONDICIÓN', 'CONDICIÓN', 'TU FÓRMULA =SUMAR.SI(...)', 'RESPUESTA CORRECTA']
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=5, column=i).value = h
    _hdr(ws3, 5, 1, 6, bg=GREEN)

    ejercicios3 = [
        (1, '¿Cuánto suma la nómina del depto. Ventas?',
            'Datos_Nomina!C3:C22', '"Ventas"',
            '=SUMAR.SI(Datos_Nomina!$C$3:$C$22,"Ventas",Datos_Nomina!$F$3:$F$22)',
            12150000, True),
        (2, '¿Cuánto suman los salarios de Barranquilla?',
            'Datos_Nomina!D3:D22', '"Barranquilla"',
            '=SUMAR.SI(Datos_Nomina!$D$3:$D$22,"Barranquilla",Datos_Nomina!$F$3:$F$22)',
            40800000, True),
        (3, '¿Cuánto suman los salarios de Logística?',
            'Datos_Nomina!C3:C22', '"Logística"',
            '', 8100000, False),
        (4, '¿Cuánto suman los salarios de Contabilidad?',
            'Datos_Nomina!C3:C22', '"Contabilidad"',
            '', 7850000, False),
        (5, '¿Cuánto suma la nómina de Bogotá?',
            'Datos_Nomina!D3:D22', '"Bogotá"',
            '', 16400000, False),
        (6, '¿Cuánto suman los bonos de empleados Senior?',
            'Datos_Nomina!E3:E22', '"Senior"',
            '', 3330000, False),
        (7, '¿Cuánto suman salarios mayores a $3,000,000?',
            'Datos_Nomina!F3:F22', '">3000000"',
            '', 33900000, False),
    ]

    for r, (num, preg, rng, cond, formula, resp, resuelto) in enumerate(ejercicios3, 6):
        _cell(ws3, r, 1, num, bold=True, align='center', color=GREEN)
        _cell(ws3, r, 2, preg, size=9, wrap=True)
        ws3.row_dimensions[r].height = 30
        _cell(ws3, r, 3, rng.replace('Datos_Nomina!', ''), color='888888', size=8)
        _cell(ws3, r, 4, cond, color='0070C0', size=9, bold=True)
        if resuelto:
            _cell(ws3, r, 5, formula, color='0070C0', fill=GREEN_LIGHT, size=9)
        else:
            _cell(ws3, r, 5, '← Escribe tu fórmula SUMAR.SI aquí', color='AAAAAA', fill=YELLOW_FILL, italic=True, size=9)
        cell_resp = ws3.cell(row=r, column=6, value=resp)
        cell_resp.number_format = '"$"#,##0'
        cell_resp.font = Font(bold=True, color=GREEN, size=11)
        cell_resp.fill = _fill(GREEN_LIGHT)
        cell_resp.border = _border()
        cell_resp.alignment = Alignment(horizontal='right', vertical='center')

    _widths(ws3, [4, 36, 25, 18, 50, 18])

    # ── HOJA 4: CONTAR_SI ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet('CONTAR_SI')
    _title_row(ws4, 1, 'CONTAR.SI — Cuenta cuántas celdas cumplen una condición', 6)
    ws4.merge_cells('A1:F1')

    ws4.cell(row=2, column=1).value = 'SINTAXIS: =CONTAR.SI( rango , "condición" )   ←  Solo 2 argumentos. No hay rango de suma.'
    ws4.cell(row=2, column=1).font = Font(bold=True, size=10, color='0070C0')
    ws4.merge_cells('A2:F2')

    ws4.cell(row=3, column=1).value = 'Operadores válidos: "texto" | ">número" | "<número" | ">=número" | "<>texto" (diferente) | "*parcial*" (comodín)'
    ws4.cell(row=3, column=1).font = Font(size=9, color='555555', italic=True)
    ws4.merge_cells('A3:F3')

    headers4 = ['#', 'PREGUNTA', 'RANGO', 'CONDICIÓN', 'TU FÓRMULA =CONTAR.SI(...)', 'RESPUESTA CORRECTA']
    for i, h in enumerate(headers4, 1):
        ws4.cell(row=5, column=i).value = h
    _hdr(ws4, 5, 1, 6, bg='00A070')

    ejercicios4 = [
        (1, '¿Cuántos empleados son del depto. Ventas?',
            'Datos_Nomina!C3:C22', '"Ventas"',
            '=CONTAR.SI(Datos_Nomina!$C$3:$C$22,"Ventas")', 5, True),
        (2, '¿Cuántos empleados son de Barranquilla?',
            'Datos_Nomina!D3:D22', '"Barranquilla"',
            '=CONTAR.SI(Datos_Nomina!$D$3:$D$22,"Barranquilla")', 13, True),
        (3, '¿Cuántos empleados son Junior?',
            'Datos_Nomina!E3:E22', '"Junior"',
            '', 5, False),
        (4, '¿Cuántos empleados ganan más de $3,000,000?',
            'Datos_Nomina!F3:F22', '">3000000"',
            '', 7, False),
        (5, '¿Cuántos empleados son de Bogotá?',
            'Datos_Nomina!D3:D22', '"Bogotá"',
            '', 5, False),
        (6, '¿Cuántos empleados tienen bono mayor a $0?',
            'Datos_Nomina!G3:G22', '">0"',
            '', 9, False),
        (7, '¿Cuántos empleados son de Contabilidad o RRHH? (usa dos CONTAR.SI)',
            'Datos_Nomina!C3:C22', '"Contab." / "RRHH"',
            '', 5, False),
    ]

    for r, (num, preg, rng, cond, formula, resp, resuelto) in enumerate(ejercicios4, 6):
        _cell(ws4, r, 1, num, bold=True, align='center', color=GREEN)
        _cell(ws4, r, 2, preg, size=9, wrap=True)
        ws4.row_dimensions[r].height = 30
        _cell(ws4, r, 3, rng.replace('Datos_Nomina!', ''), color='888888', size=8)
        _cell(ws4, r, 4, cond, color='0070C0', size=9, bold=True)
        if resuelto:
            _cell(ws4, r, 5, formula, color='0070C0', fill=GREEN_LIGHT, size=9)
        else:
            _cell(ws4, r, 5, '← Escribe tu fórmula CONTAR.SI aquí', color='AAAAAA', fill=YELLOW_FILL, italic=True, size=9)
        cell_resp = ws4.cell(row=r, column=6, value=resp)
        cell_resp.font = Font(bold=True, color=GREEN, size=14)
        cell_resp.fill = _fill(GREEN_LIGHT)
        cell_resp.border = _border()
        cell_resp.alignment = Alignment(horizontal='center', vertical='center')

    _widths(ws4, [4, 40, 25, 20, 50, 18])

    # ── HOJA 5: PROMEDIO_SI ───────────────────────────────────────────────────
    ws5 = wb.create_sheet('PROMEDIO_SI')
    _title_row(ws5, 1, 'PROMEDIO.SI — Promedio de valores que cumplen una condición', 6)
    ws5.merge_cells('A1:F1')

    ws5.cell(row=2, column=1).value = 'SINTAXIS: =PROMEDIO.SI( rango_condición , "condición" , rango_valores_a_promediar )'
    ws5.cell(row=2, column=1).font = Font(bold=True, size=10, color='0070C0')
    ws5.merge_cells('A2:F2')

    ws5.cell(row=3, column=1).value = 'Misma estructura que SUMAR.SI. Si rango_condición = rango_valores, omite el tercer argumento.'
    ws5.cell(row=3, column=1).font = Font(size=9, color='555555', italic=True)
    ws5.merge_cells('A3:F3')

    headers5 = ['#', 'PREGUNTA', 'RANGO CONDICIÓN', 'CONDICIÓN', 'TU FÓRMULA =PROMEDIO.SI(...)', 'RESPUESTA CORRECTA']
    for i, h in enumerate(headers5, 1):
        ws5.cell(row=5, column=i).value = h
    _hdr(ws5, 5, 1, 6, bg=BLUE, fg='1a1a2a')

    ejercicios5 = [
        (1, '¿Cuál es el salario promedio del depto. Ventas?',
            'Datos_Nomina!C3:C22', '"Ventas"',
            '=PROMEDIO.SI(Datos_Nomina!$C$3:$C$22,"Ventas",Datos_Nomina!$F$3:$F$22)', 2430000, True),
        (2, '¿Cuál es el salario promedio en Barranquilla?',
            'Datos_Nomina!D3:D22', '"Barranquilla"',
            '=PROMEDIO.SI(Datos_Nomina!$D$3:$D$22,"Barranquilla",Datos_Nomina!$F$3:$F$22)', 3138462, True),
        (3, '¿Cuál es el salario promedio de empleados Senior?',
            'Datos_Nomina!E3:E22', '"Senior"',
            '', 4162500, False),
        (4, '¿Cuál es el salario promedio de Contabilidad?',
            'Datos_Nomina!C3:C22', '"Contabilidad"',
            '', 2616667, False),
        (5, '¿Cuál es el promedio de bonos (solo de quien tiene bono)?',
            'Datos_Nomina!G3:G22', '">0"',
            '', 370000, False),
        (6, '¿Cuál es el salario promedio en Bogotá?',
            'Datos_Nomina!D3:D22', '"Bogotá"',
            '', 3280000, False),
    ]

    for r, (num, preg, rng, cond, formula, resp, resuelto) in enumerate(ejercicios5, 6):
        _cell(ws5, r, 1, num, bold=True, align='center', color=BLUE)
        _cell(ws5, r, 2, preg, size=9, wrap=True)
        ws5.row_dimensions[r].height = 30
        _cell(ws5, r, 3, rng.replace('Datos_Nomina!', ''), color='888888', size=8)
        _cell(ws5, r, 4, cond, color='0070C0', size=9, bold=True)
        if resuelto:
            _cell(ws5, r, 5, formula, color='0070C0', fill=GREEN_LIGHT, size=9)
        else:
            _cell(ws5, r, 5, '← Escribe tu fórmula PROMEDIO.SI aquí', color='AAAAAA', fill=YELLOW_FILL, italic=True, size=9)
        cell_resp = ws5.cell(row=r, column=6, value=resp)
        cell_resp.number_format = '"$"#,##0'
        cell_resp.font = Font(bold=True, color=BLUE, size=11)
        cell_resp.fill = _fill('E8F0FF')
        cell_resp.border = _border()
        cell_resp.alignment = Alignment(horizontal='right', vertical='center')

    _widths(ws5, [4, 42, 25, 18, 55, 18])

    # ── HOJA 6: RETO_CONJUNTO ─────────────────────────────────────────────────
    ws6 = wb.create_sheet('RETO_CONJUNTO')
    _title_row(ws6, 1, 'RETO — SUMAR.SI.CONJUNTO: Múltiples condiciones simultáneas', 6)
    ws6.merge_cells('A1:F1')

    ws6.cell(row=2, column=1).value = '¡ATENCIÓN AL ORDEN! En SUMAR.SI.CONJUNTO el rango_suma va PRIMERO (al contrario de SUMAR.SI)'
    ws6.cell(row=2, column=1).font = Font(bold=True, size=10, color=RED)
    ws6.merge_cells('A2:F2')

    ws6.cell(row=3, column=1).value = 'SINTAXIS: =SUMAR.SI.CONJUNTO( rango_suma , rango_crit1 , crit1 , rango_crit2 , crit2 , ... )'
    ws6.cell(row=3, column=1).font = Font(size=10, color='0070C0', bold=True)
    ws6.merge_cells('A3:F3')

    ws6.cell(row=4, column=1).value = 'Puedes agregar hasta 127 pares (rango_criterio, criterio). Se suman solo las filas que cumplen TODAS las condiciones.'
    ws6.cell(row=4, column=1).font = Font(size=9, color='555555', italic=True)
    ws6.merge_cells('A4:F4')

    headers6 = ['#', 'PREGUNTA (2 condiciones)', 'CONDICIÓN 1', 'CONDICIÓN 2', 'TU FÓRMULA', 'RESPUESTA']
    for i, h in enumerate(headers6, 1):
        ws6.cell(row=6, column=i).value = h
    _hdr(ws6, 6, 1, 6, bg=AMBER, fg='1a1000')

    ejercicios6 = [
        (1, '¿Cuánto suma la nómina de Ventas en Barranquilla?',
            'C="Ventas"', 'D="Barranquilla"',
            '=SUMAR.SI.CONJUNTO(Datos_Nomina!$F$3:$F$22,Datos_Nomina!$C$3:$C$22,"Ventas",Datos_Nomina!$D$3:$D$22,"Barranquilla")',
            8500000, True),
        (2, '¿Cuánto suma la nómina de empleados Senior en Barranquilla?',
            'E="Senior"', 'D="Barranquilla"',
            '=SUMAR.SI.CONJUNTO(Datos_Nomina!$F$3:$F$22,Datos_Nomina!$E$3:$E$22,"Senior",Datos_Nomina!$D$3:$D$22,"Barranquilla")',
            28350000, True),
        (3, '¿Cuánto suman los bonos de Ventas en Bogotá?',
            'C="Ventas"', 'D="Bogotá"',
            '', 300000, False),
        (4, '¿Cuánto suman los salarios de Logística en Bogotá?',
            'C="Logística"', 'D="Bogotá"',
            '', 6400000, False),
        (5, '¿Cuánto suma la nómina de Auxiliares en Barranquilla?',
            'E="Auxiliar"', 'D="Barranquilla"',
            '', 5250000, False),
        (6, '¿Cuánto suman bonos de Senior en Bogotá?',
            'E="Senior"', 'D="Bogotá"',
            '', 1080000, False),
    ]

    for r, (num, preg, c1, c2, formula, resp, resuelto) in enumerate(ejercicios6, 7):
        _cell(ws6, r, 1, num, bold=True, align='center', color=AMBER)
        _cell(ws6, r, 2, preg, size=9, wrap=True)
        ws6.row_dimensions[r].height = 35
        _cell(ws6, r, 3, c1, color='555555', size=9)
        _cell(ws6, r, 4, c2, color='555555', size=9)
        if resuelto:
            _cell(ws6, r, 5, formula, color='0070C0', fill=GREEN_LIGHT, size=8, wrap=True)
        else:
            _cell(ws6, r, 5, '← Escribe tu fórmula SUMAR.SI.CONJUNTO aquí', color='AAAAAA', fill=YELLOW_FILL, italic=True, size=9)
        cell_resp = ws6.cell(row=r, column=6, value=resp)
        cell_resp.number_format = '"$"#,##0'
        cell_resp.font = Font(bold=True, color=AMBER, size=11)
        cell_resp.fill = _fill('FFF3CD')
        cell_resp.border = _border()
        cell_resp.alignment = Alignment(horizontal='right', vertical='center')

    _widths(ws6, [4, 38, 20, 20, 55, 18])

    path = os.path.join(BASE, 'actividad_02_Funciones_Condicionales.xlsx')
    wb.save(path)
    print(f'  ✓ {os.path.relpath(path)}')


# ─────────────────────────────────────────────────────────────────────────────
# ARCHIVO 3: TAREA PARA CASA — VIDA REAL
# ─────────────────────────────────────────────────────────────────────────────
def tarea_vida_real():
    wb = openpyxl.Workbook()

    CORAL  = 'E8472A'
    TEAL   = '009688'
    ORANGE = 'FF6B35'
    CREAM  = 'FFFDE7'
    TEAL_LIGHT = 'E0F2F1'

    # ── HOJA 1: INSTRUCCIONES ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'INSTRUCCIONES'
    ws.column_dimensions['A'].width = 92

    inst = [
        (1,  'TAREA PARA LA CASA — PRESUPUESTO FAMILIAR CON EXCEL',         True,  16, TEAL),
        (2,  'DÍA 2 | Módulo 11 — Ofimática Avanzada | INTEBIEN S.A.S.',    False, 10, '888888'),
        (3,  '', False, 10, '333333'),
        (4,  'CONTEXTO — ¿Por qué esta tarea?', True, 12, TEAL),
        (5,  'Excel no es solo para el trabajo. Estas mismas funciones que usas en la empresa', False, 10, '333333'),
        (6,  'pueden ayudarte a controlar tus gastos personales, planear vacaciones, hacer listas', False, 10, '333333'),
        (7,  'de compras inteligentes, y tomar mejores decisiones financieras en tu día a día.', False, 10, '333333'),
        (8,  '', False, 10, '333333'),
        (9,  'SITUACIÓN: Eres el administrador del presupuesto de tu hogar en junio 2025.', True, 11, TEAL),
        (10, 'Tienes un registro de todos los gastos del mes y un catálogo del supermercado.', False, 10, '333333'),
        (11, '', False, 10, '333333'),
        (12, 'ESTRUCTURA DE ESTE ARCHIVO', True, 12, TEAL),
        (13, '  Hoja 2 — Gastos_Junio       : 35 registros de gastos reales del mes (ya está lleno)', False, 10, '333333'),
        (14, '  Hoja 3 — Catalogo_Super     : Precios de 20 productos del supermercado', False, 10, '333333'),
        (15, '  Hoja 4 — Lista_Compras      : PRÁCTICA INDICE+COINCIDIR — buscar precios', False, 10, '333333'),
        (16, '  Hoja 5 — Analisis_Presupuesto: PRÁCTICA CONDICIONALES — analizar gastos', False, 10, '333333'),
        (17, '', False, 10, '333333'),
        (18, 'LO QUE DEBES HACER', True, 12, TEAL),
        (19, '', False, 10, '333333'),
        (20, '  PARTE A — Lista de Compras (Hoja 4)', True, 11, ORANGE),
        (21, '  1. Abre la hoja "Lista_Compras"', False, 10, '333333'),
        (22, '  2. En la columna D, usa INDICE+COINCIDIR para traer el precio de cada producto', False, 10, '333333'),
        (23, '     desde el catálogo del supermercado (Hoja 3)', False, 10, '333333'),
        (24, '  3. En la columna E, calcula el subtotal (cantidad × precio)', False, 10, '333333'),
        (25, '  4. Calcula el total de la compra', False, 10, '333333'),
        (26, '', False, 10, '333333'),
        (27, '  PARTE B — Análisis del Presupuesto (Hoja 5)', True, 11, ORANGE),
        (28, '  1. Abre la hoja "Analisis_Presupuesto"', False, 10, '333333'),
        (29, '  2. Usa SUMAR.SI para calcular cuánto gastaste en cada categoría del mes', False, 10, '333333'),
        (30, '  3. Usa CONTAR.SI para saber cuántas transacciones tuvo cada categoría', False, 10, '333333'),
        (31, '  4. Usa PROMEDIO.SI para saber el gasto promedio por transacción en cada categoría', False, 10, '333333'),
        (32, '  5. Responde las 3 preguntas de análisis al final de la hoja', False, 10, '333333'),
        (33, '', False, 10, '333333'),
        (34, 'ENTREGA: Guarda como  Tarea_D2_TuNombre.xlsx', True, 11, CORAL),
        (35, '', False, 10, '333333'),
        (36, 'PISTAS CLAVE:', True, 11, TEAL),
        (37, '  INDICE+COINCIDIR para buscar precio por nombre de producto:', False, 10, '333333'),
        (38, '  =IFERROR(INDICE(Catalogo_Super!$C$3:$C$22,COINCIDIR(B3,Catalogo_Super!$A$3:$A$22,0)),0)', False, 10, '0070C0'),
        (39, '  SUMAR.SI para total de una categoría:', False, 10, '333333'),
        (40, '  =SUMAR.SI(Gastos_Junio!$C$3:$C$37,"Alimentación",Gastos_Junio!$E$3:$E$37)', False, 10, '0070C0'),
    ]
    for row_i, (row, text, bold, size, color) in enumerate(inst):
        c = ws.cell(row=row, column=1, value=text)
        c.font = Font(bold=bold, size=size, color=color)

    # ── HOJA 2: GASTOS_JUNIO ──────────────────────────────────────────────────
    ws2 = wb.create_sheet('Gastos_Junio')
    _title_row(ws2, 1, 'REGISTRO DE GASTOS — JUNIO 2025 — HOGAR FAMILIAR', 6)
    ws2.merge_cells('A1:F1')
    ws2.cell(row=1, column=1).font = Font(bold=True, size=13, color=TEAL)

    headers2 = ['FECHA', 'DESCRIPCIÓN', 'CATEGORÍA', 'FORMA_PAGO', 'MONTO', 'NECESIDAD']
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=2, column=i).value = h
    _hdr(ws2, 2, 1, 6, bg=TEAL)

    gastos = [
        ('2025-06-01', 'Arriendo apartamento',            'Vivienda',         'Transferencia', 1200000, 'Obligatorio'),
        ('2025-06-01', 'Mercado semanal semana 1',         'Alimentación',     'Débito',         280000, 'Obligatorio'),
        ('2025-06-02', 'Recarga bus mensual',              'Transporte',       'Efectivo',        75000, 'Obligatorio'),
        ('2025-06-03', 'Netflix',                          'Entretenimiento',  'Crédito',         22900, 'Opcional'),
        ('2025-06-04', 'Desayuno panadería',               'Alimentación',     'Efectivo',        15000, 'Opcional'),
        ('2025-06-05', 'Gasolina carro',                   'Transporte',       'Crédito',         80000, 'Obligatorio'),
        ('2025-06-06', 'Gym mensualidad',                  'Salud',            'Débito',          60000, 'Opcional'),
        ('2025-06-07', 'Almuerzo restaurante con familia', 'Alimentación',     'Crédito',         95000, 'Opcional'),
        ('2025-06-08', 'Servicio de energía',              'Servicios',        'PSE',            145000, 'Obligatorio'),
        ('2025-06-08', 'Servicio de agua',                 'Servicios',        'PSE',             65000, 'Obligatorio'),
        ('2025-06-09', 'Internet hogar',                   'Servicios',        'Débito',          75000, 'Obligatorio'),
        ('2025-06-10', 'Útiles escolares hijos',           'Educación',        'Efectivo',        85000, 'Obligatorio'),
        ('2025-06-11', 'Mercado semanal semana 2',         'Alimentación',     'Débito',         255000, 'Obligatorio'),
        ('2025-06-12', 'Médico pediatra',                  'Salud',            'Efectivo',        60000, 'Obligatorio'),
        ('2025-06-13', 'Cine en familia',                  'Entretenimiento',  'Crédito',         85000, 'Opcional'),
        ('2025-06-14', 'Ropa deportiva hijo',              'Ropa',             'Crédito',         95000, 'Opcional'),
        ('2025-06-15', 'Taxi a reunión de trabajo',        'Transporte',       'App',             22000, 'Obligatorio'),
        ('2025-06-16', 'Spotify',                          'Entretenimiento',  'Crédito',         11900, 'Opcional'),
        ('2025-06-17', 'Domicilio almuerzo',               'Alimentación',     'App',             35000, 'Opcional'),
        ('2025-06-18', 'Peluquería',                       'Cuidado Personal', 'Efectivo',        40000, 'Opcional'),
        ('2025-06-19', 'Medicamentos farmacia',            'Salud',            'Débito',          32000, 'Obligatorio'),
        ('2025-06-20', 'Gas domiciliario',                 'Servicios',        'PSE',             38000, 'Obligatorio'),
        ('2025-06-21', 'Mercado semanal semana 3',         'Alimentación',     'Débito',         270000, 'Obligatorio'),
        ('2025-06-22', 'Parqueadero oficina x mes',        'Transporte',       'Transferencia',   80000, 'Obligatorio'),
        ('2025-06-23', 'Libro para el trabajo',            'Educación',        'Amazon',          48000, 'Opcional'),
        ('2025-06-24', 'Cumpleaños — regalo amigo',        'Entretenimiento',  'Efectivo',        65000, 'Opcional'),
        ('2025-06-25', 'Desayuno empanadas',               'Alimentación',     'Efectivo',         8000, 'Opcional'),
        ('2025-06-26', 'Revisión carros mecánica',         'Transporte',       'Transferencia',  180000, 'Obligatorio'),
        ('2025-06-27', 'Supermercado aseo y limpieza',     'Hogar',            'Débito',          55000, 'Obligatorio'),
        ('2025-06-28', 'Teléfono celular plan',            'Servicios',        'Débito',          49000, 'Obligatorio'),
        ('2025-06-29', 'Mercado semanal semana 4',         'Alimentación',     'Débito',         245000, 'Obligatorio'),
        ('2025-06-30', 'Crédito / cuota del mes',          'Deudas',           'Débito',         350000, 'Obligatorio'),
        ('2025-06-30', 'Salida a parque de diversiones',   'Entretenimiento',  'Efectivo',       120000, 'Opcional'),
        ('2025-06-30', 'Barbería',                         'Cuidado Personal', 'Efectivo',        20000, 'Opcional'),
        ('2025-06-30', 'Papelería escolar final mes',      'Educación',        'Efectivo',        35000, 'Obligatorio'),
    ]

    alts = [TEAL_LIGHT, WHITE]
    for r, row in enumerate(gastos, 3):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c == 5:
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = _fill(alts[r % 2])
            cell.border = _border()

    ws2.cell(row=39, column=4).value = 'TOTAL MES:'
    ws2.cell(row=39, column=4).font = Font(bold=True, size=11)
    ws2.cell(row=39, column=4).border = _border()
    ws2.cell(row=39, column=5).value = '=SUM(E3:E37)'
    ws2.cell(row=39, column=5).number_format = '"$"#,##0'
    ws2.cell(row=39, column=5).font = Font(bold=True, size=13, color=TEAL)
    ws2.cell(row=39, column=5).fill = _fill(TEAL_LIGHT)
    ws2.cell(row=39, column=5).border = _border()

    _widths(ws2, [13, 38, 17, 15, 14, 13])
    _freeze(ws2)
    ws2.auto_filter.ref = 'A2:F2'

    # ── HOJA 3: CATALOGO_SUPER ────────────────────────────────────────────────
    ws3 = wb.create_sheet('Catalogo_Super')
    _title_row(ws3, 1, 'CATÁLOGO SUPERMERCADO — PRECIOS REFERENCIA JUNIO 2025', 5)
    ws3.merge_cells('A1:E1')
    ws3.cell(row=1, column=1).font = Font(bold=True, size=13, color=ORANGE)

    headers3 = ['PRODUCTO', 'UNIDAD', 'PRECIO_UNITARIO', 'CATEGORÍA', 'PASILLO']
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=2, column=i).value = h
    _hdr(ws3, 2, 1, 5, bg=ORANGE, fg='1a0a00')

    productos = [
        ('Arroz Diana x 500g',        '500g',   3800,  'Granos',        'A1'),
        ('Aceite Gourmet x 1L',       '1 Lt',   9500,  'Aceites',       'A2'),
        ('Leche Alquería x 1L',       '1 Lt',   4200,  'Lácteos',       'B1'),
        ('Huevos AA x 12 und',        '12 und', 13500, 'Proteínas',     'B2'),
        ('Pan tajado Bimbo',           '1 bolsa', 5600, 'Panadería',     'C1'),
        ('Pollo entero x kg',         'kg',     12000, 'Proteínas',     'B3'),
        ('Papa pastusa x kg',         'kg',      2800, 'Verduras',      'D1'),
        ('Cebolla cabezona x kg',     'kg',      3500, 'Verduras',      'D2'),
        ('Tomate chonto x kg',        'kg',      4200, 'Verduras',      'D3'),
        ('Zanahoria x 500g',          '500g',    3200, 'Verduras',      'D4'),
        ('Pasta Doria x 500g',        '500g',    4500, 'Granos',        'A3'),
        ('Atún Van Camps x lata',     'lata',    5800, 'Proteínas',     'B4'),
        ('Café Sello Rojo x 500g',    '500g',   22000, 'Bebidas',       'E1'),
        ('Azúcar Manuelita x kg',     'kg',      4200, 'Dulces',        'A4'),
        ('Sal Refisal x kg',          'kg',      2100, 'Condimentos',   'A5'),
        ('Detergente Ariel x 1kg',    'kg',     18500, 'Aseo',          'F1'),
        ('Jabón Palmolive x 3 bar',   '3 barras', 7800, 'Aseo',         'F2'),
        ('Papel higiénico Scott x 12','12 rollos', 28000, 'Aseo',       'F3'),
        ('Shampoo Head&Shoulders',    '1 bot',  18500, 'Higiene',       'G1'),
        ('Gaseosa Postobon x 2L',     '2 Lt',    6500, 'Bebidas',       'E2'),
    ]

    alts = ['FFF3E0', WHITE]
    for r, row in enumerate(productos, 3):
        for c, val in enumerate(row, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            if c == 3:
                cell.number_format = '"$"#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.fill = _fill(alts[r % 2])
            cell.border = _border()

    ws3.cell(row=24, column=1).value = '← No modificar. Esta tabla es la fuente para INDICE+COINCIDIR en la hoja Lista_Compras'
    ws3.cell(row=24, column=1).font = Font(italic=True, color='888888', size=9)
    ws3.merge_cells('A24:E24')

    _widths(ws3, [30, 12, 18, 14, 10])
    _freeze(ws3)

    # ── HOJA 4: LISTA_COMPRAS ─────────────────────────────────────────────────
    ws4 = wb.create_sheet('Lista_Compras')
    _title_row(ws4, 1, 'LISTA DE COMPRAS — USA INDICE+COINCIDIR PARA BUSCAR LOS PRECIOS', 6)
    ws4.merge_cells('A1:F1')
    ws4.cell(row=1, column=1).font = Font(bold=True, size=13, color=ORANGE)

    ws4.cell(row=2, column=1).value = 'TAREA: En la columna D, usa INDICE+COINCIDIR para traer el precio desde el Catalogo_Super'
    ws4.cell(row=2, column=1).font = Font(bold=True, size=10, color='333333')
    ws4.merge_cells('A2:F2')

    ws4.cell(row=3, column=1).value = 'FÓRMULA A USAR:  =IFERROR(INDICE(Catalogo_Super!$C$3:$C$22, COINCIDIR(B4, Catalogo_Super!$A$3:$A$22, 0)), 0)'
    ws4.cell(row=3, column=1).font = Font(size=10, color='0070C0', bold=True)
    ws4.merge_cells('A3:F3')

    headers4 = ['#', 'PRODUCTO', 'CANTIDAD', 'PRECIO UNITARIO (usa fórmula)', 'SUBTOTAL (Cant × Precio)', 'VERIFICACIÓN']
    for i, h in enumerate(headers4, 1):
        ws4.cell(row=5, column=i).value = h
    _hdr(ws4, 5, 1, 6, bg=ORANGE, fg='1a0a00')

    lista_compras = [
        ('Arroz Diana x 500g',      3, 3800),
        ('Aceite Gourmet x 1L',     2, 9500),
        ('Leche Alquería x 1L',     6, 4200),
        ('Huevos AA x 12 und',      2, 13500),
        ('Pan tajado Bimbo',         1, 5600),
        ('Pollo entero x kg',       2, 12000),
        ('Papa pastusa x kg',       3, 2800),
        ('Tomate chonto x kg',      1, 4200),
        ('Pasta Doria x 500g',      4, 4500),
        ('Atún Van Camps x lata',   6, 5800),
        ('Café Sello Rojo x 500g',  1, 22000),
        ('Azúcar Manuelita x kg',   2, 4200),
        ('Detergente Ariel x 1kg',  1, 18500),
        ('Papel higiénico Scott x 12', 1, 28000),
        ('Gaseosa Postobon x 2L',   2, 6500),
    ]

    alts = ['FFF3E0', WHITE]
    for r, (prod, cant, precio) in enumerate(lista_compras, 6):
        _cell(ws4, r, 1, r - 5, align='center', bold=True, color=ORANGE)
        _cell(ws4, r, 2, prod, fill=alts[r % 2])
        _cell(ws4, r, 3, cant, align='center', fill=alts[r % 2])
        # Columna D: celda amarilla para fórmula (primeras 2 están resueltas)
        if r <= 7:
            formula_ic = f'=IFERROR(INDICE(Catalogo_Super!$C$3:$C$22,COINCIDIR(B{r},Catalogo_Super!$A$3:$A$22,0)),0)'
            cell_d = ws4.cell(row=r, column=4, value=formula_ic)
            cell_d.number_format = '"$"#,##0'
            cell_d.font = Font(color='0070C0', size=10)
            cell_d.fill = _fill(GREEN_LIGHT)
            cell_d.border = _border()
            cell_d.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell_d = ws4.cell(row=r, column=4, value=None)
            cell_d.fill = _fill(YELLOW_FILL)
            cell_d.border = _border()
            cell_d.font = Font(size=10, italic=True, color='AAAAAA')
            cell_d.value = '← TU FÓRMULA INDICE+COINCIDIR'
            cell_d.alignment = Alignment(horizontal='left', vertical='center')
        # Columna E: subtotal (también vacío para estudiante)
        if r <= 7:
            cell_e = ws4.cell(row=r, column=5, value=f'=C{r}*D{r}')
            cell_e.number_format = '"$"#,##0'
            cell_e.font = Font(bold=True, color=TEAL)
            cell_e.fill = _fill(TEAL_LIGHT)
            cell_e.border = _border()
            cell_e.alignment = Alignment(horizontal='right', vertical='center')
        else:
            cell_e = ws4.cell(row=r, column=5, value=None)
            cell_e.fill = _fill(YELLOW_FILL)
            cell_e.border = _border()
            cell_e.font = Font(size=10, italic=True, color='AAAAAA')
            cell_e.value = f'← =C{r}*D{r}'
            cell_e.alignment = Alignment(horizontal='left', vertical='center')
        # Columna F: verificación (valor correcto)
        cell_f = ws4.cell(row=r, column=6, value=cant * precio)
        cell_f.number_format = '"$"#,##0'
        cell_f.font = Font(size=9, color='AAAAAA', italic=True)
        cell_f.border = _border()
        cell_f.alignment = Alignment(horizontal='right', vertical='center')

    # Total
    ws4.cell(row=22, column=4).value = 'TOTAL DE LA COMPRA:'
    ws4.cell(row=22, column=4).font = Font(bold=True, size=11)
    ws4.cell(row=22, column=4).border = _border()
    ws4.cell(row=22, column=5).value = '← =SUMA(E6:E20)'
    ws4.cell(row=22, column=5).fill = _fill(YELLOW_FILL)
    ws4.cell(row=22, column=5).font = Font(bold=True, size=14, color=ORANGE, italic=True)
    ws4.cell(row=22, column=5).border = _border()
    ws4.cell(row=22, column=6).value = 'Respuesta: $368,400'
    ws4.cell(row=22, column=6).font = Font(size=9, italic=True, color='AAAAAA')
    ws4.cell(row=22, column=6).border = _border()

    _widths(ws4, [4, 35, 12, 35, 32, 16])

    # ── HOJA 5: ANALISIS_PRESUPUESTO ──────────────────────────────────────────
    ws5 = wb.create_sheet('Analisis_Presupuesto')
    _title_row(ws5, 1, 'ANÁLISIS DEL PRESUPUESTO FAMILIAR — JUNIO 2025', 6)
    ws5.merge_cells('A1:F1')
    ws5.cell(row=1, column=1).font = Font(bold=True, size=13, color=TEAL)

    ws5.cell(row=2, column=1).value = 'Usa SUMAR.SI, CONTAR.SI y PROMEDIO.SI con los datos de la hoja "Gastos_Junio" para completar este análisis'
    ws5.cell(row=2, column=1).font = Font(bold=True, size=10, color='333333')
    ws5.merge_cells('A2:F2')

    # Panel principal
    headers5 = ['CATEGORÍA', 'TOTAL GASTADO (SUMAR.SI)', 'N° GASTOS (CONTAR.SI)', 'PROM. POR GASTO (PROMEDIO.SI)', '% DEL TOTAL', 'EVALUACIÓN']
    for i, h in enumerate(headers5, 1):
        ws5.cell(row=4, column=i).value = h
    _hdr(ws5, 4, 1, 6, bg=TEAL)

    categorias = [
        ('Alimentación',     958000),
        ('Transporte',       437000),
        ('Servicios',        372000),
        ('Entretenimiento',  304800),
        ('Salud',            152000),
        ('Vivienda',        1200000),
        ('Educación',        168000),
        ('Ropa',              95000),
        ('Cuidado Personal',  60000),
        ('Hogar',             55000),
        ('Deudas',           350000),
    ]

    total_correcto = sum(v for _, v in categorias)

    for r, (cat, total_esp) in enumerate(categorias, 5):
        _cell(ws5, r, 1, cat, bold=True)
        # SUMAR.SI
        cell_b = ws5.cell(row=r, column=2)
        cell_b.value = f'← =SUMAR.SI(Gastos_Junio!$C:$C,"{cat}",Gastos_Junio!$E:$E)'
        cell_b.fill = _fill(YELLOW_FILL)
        cell_b.font = Font(size=9, italic=True, color='AAAAAA')
        cell_b.border = _border()
        cell_b.number_format = '"$"#,##0'
        # CONTAR.SI
        cell_c = ws5.cell(row=r, column=3)
        cell_c.value = f'← =CONTAR.SI(Gastos_Junio!$C:$C,"{cat}")'
        cell_c.fill = _fill(YELLOW_FILL)
        cell_c.font = Font(size=9, italic=True, color='AAAAAA')
        cell_c.border = _border()
        # PROMEDIO.SI
        cell_d = ws5.cell(row=r, column=4)
        cell_d.value = f'← =IFERROR(PROMEDIO.SI(Gastos_Junio!$C:$C,"{cat}",Gastos_Junio!$E:$E),0)'
        cell_d.fill = _fill(YELLOW_FILL)
        cell_d.font = Font(size=9, italic=True, color='AAAAAA')
        cell_d.border = _border()
        cell_d.number_format = '"$"#,##0'
        # % del total (necesita B5:B15 total)
        cell_e = ws5.cell(row=r, column=5)
        cell_e.value = '← =B{}/SUMA($B$5:$B$15)'.format(r)
        cell_e.fill = _fill(YELLOW_FILL)
        cell_e.font = Font(size=9, italic=True, color='AAAAAA')
        cell_e.border = _border()
        cell_e.number_format = '0.0%'
        # Verificación
        cell_f = ws5.cell(row=r, column=6, value=total_esp)
        cell_f.number_format = '"$"#,##0'
        cell_f.font = Font(size=9, italic=True, color='AAAAAA')
        cell_f.border = _border()
        cell_f.alignment = Alignment(horizontal='right', vertical='center')

    # Fila total
    total_row = 16
    ws5.cell(row=total_row, column=1).value = 'TOTAL GASTO MENSUAL'
    ws5.cell(row=total_row, column=1).font = Font(bold=True, size=11)
    ws5.cell(row=total_row, column=1).border = _border()
    ws5.cell(row=total_row, column=2).value = '=SUMA(B5:B15)'
    ws5.cell(row=total_row, column=2).number_format = '"$"#,##0'
    ws5.cell(row=total_row, column=2).font = Font(bold=True, size=13, color=TEAL)
    ws5.cell(row=total_row, column=2).fill = _fill(TEAL_LIGHT)
    ws5.cell(row=total_row, column=2).border = _border()
    ws5.cell(row=total_row, column=6).value = total_correcto
    ws5.cell(row=total_row, column=6).number_format = '"$"#,##0'
    ws5.cell(row=total_row, column=6).font = Font(size=9, italic=True, color='AAAAAA')
    ws5.cell(row=total_row, column=6).border = _border()

    # Preguntas de análisis
    ws5.cell(row=18, column=1).value = 'PREGUNTAS PARA REFLEXIONAR:'
    ws5.cell(row=18, column=1).font = Font(bold=True, size=11, color=TEAL)
    ws5.merge_cells('A18:F18')

    preguntas5 = [
        (19, '1. ¿En qué categoría gastaste más dinero este mes? ¿Es un gasto obligatorio o podría reducirse?'),
        (20, '2. ¿Cuánto representan los gastos "opcionales" (entretenimiento + ropa + cuidado) del total?'),
        (21, '3. Si quisieras ahorrar $200,000 el próximo mes, ¿en qué categorías recortarías? ¿Cómo?'),
    ]
    for row_n, preg in preguntas5:
        ws5.cell(row=row_n, column=1).value = preg
        ws5.cell(row=row_n, column=1).font = Font(size=10, bold=True, color='333333')
        ws5.merge_cells(f'A{row_n}:C{row_n}')
        ws5.cell(row=row_n, column=4).value = '← Escribe tu respuesta aquí'
        ws5.cell(row=row_n, column=4).fill = _fill(YELLOW_FILL)
        ws5.cell(row=row_n, column=4).font = Font(size=10, italic=True, color='AAAAAA')
        ws5.cell(row=row_n, column=4).border = _border()
        ws5.merge_cells(f'D{row_n}:F{row_n}')
        ws5.row_dimensions[row_n].height = 28

    _widths(ws5, [20, 38, 24, 38, 14, 16])

    path = os.path.join(BASE, 'tarea_casa_Vida_Real.xlsx')
    wb.save(path)
    print(f'  ✓ {os.path.relpath(path)}')


if __name__ == '__main__':
    print('Generando archivos actividades Día 2...')
    actividad_indice_coincidir()
    actividad_condicionales()
    tarea_vida_real()
    print('\n✓ Todos los archivos generados en dia02/')
