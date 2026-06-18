#!/usr/bin/env python3
"""Genera archivos Excel de ejemplo para cada día del módulo 11."""
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.series import DataPoint
import os
import random

BASE = os.path.dirname(os.path.abspath(__file__))

# ── helpers ──────────────────────────────────────────────────────────────────
def thin():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(color='1AB87A'):
    return PatternFill('solid', fgColor=color)

def style_header(ws, row, cols, bg='1AB87A', fg='FFFFFF', bold=True):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = Font(bold=bold, color=fg, size=10)
        cell.fill = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin()

def style_data_row(ws, row, cols, bg=None):
    fill = PatternFill('solid', fgColor=bg) if bg else None
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        if fill:
            cell.fill = fill
        cell.border = thin()
        cell.alignment = Alignment(vertical='center')

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell='A2'):
    ws.freeze_panes = cell

def autofilter(ws, ref):
    ws.auto_filter.ref = ref

def save(wb, path):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f'  ✓ {os.path.relpath(path, BASE)}')

# ── Día 01: BUSCARV / BUSCARH ────────────────────────────────────────────────
def dia01():
    wb = openpyxl.Workbook()

    # --- Hoja 1: Catálogo de Productos ---
    ws = wb.active
    ws.title = 'Catalogo_Productos'
    ws['A1'] = 'BASE DE DATOS — CATÁLOGO DE PRODUCTOS — ADMIN CARIBE S.A.S.'
    ws['A1'].font = Font(bold=True, size=12, color='1AB87A')
    ws.merge_cells('A1:F1')

    headers = ['COD_PRODUCTO', 'NOMBRE_PRODUCTO', 'CATEGORIA', 'PRECIO_UNIT', 'STOCK', 'PROVEEDOR']
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i).value = h
    style_header(ws, 2, 6, '1AB87A')

    productos = [
        ('P001','Resma de Papel Carta','Papelería',18500,150,'Papelería Sur'),
        ('P002','Bolígrafo Azul x12','Papelería',12000,300,'Distribuidora Norte'),
        ('P003','Carpeta AZ Oficio','Archivos',8500,200,'Papelería Sur'),
        ('P004','Tóner HP LaserJet','Tecnología',185000,40,'Tech Soluciones'),
        ('P005','Memoria USB 32GB','Tecnología',35000,80,'Tech Soluciones'),
        ('P006','Sello Automático','Oficina',45000,25,'Distribuidora Norte'),
        ('P007','Folder Manila x50','Papelería',22000,120,'Papelería Sur'),
        ('P008','Calculadora Científica','Oficina',95000,30,'Tech Soluciones'),
        ('P009','Resaltador x5 colores','Papelería',9500,250,'Distribuidora Norte'),
        ('P010','Block Cuadriculado','Papelería',5500,400,'Papelería Sur'),
        ('P011','Tijeras Oficina','Herramientas',12500,60,'Distribuidora Norte'),
        ('P012','Grapadora Metálica','Herramientas',38000,45,'Papelería Sur'),
        ('P013','Perforadora 2 huecos','Herramientas',28000,55,'Distribuidora Norte'),
        ('P014','Rotuladores x8','Papelería',18000,180,'Papelería Sur'),
        ('P015','Archivador Vertical','Archivos',125000,20,'Muebles Caribe'),
    ]

    alts = ['F2F7FF','FFFFFF']
    for r, row in enumerate(productos, 3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 4:
                cell.number_format = '"$"#,##0'
        style_data_row(ws, r, 6, alts[r % 2])

    set_col_widths(ws, [14, 28, 14, 14, 10, 22])
    freeze(ws)
    autofilter(ws, 'A2:F2')

    # --- Hoja 2: Facturas ---
    ws2 = wb.create_sheet('Facturas_Ventas')
    ws2['A1'] = 'REGISTRO DE VENTAS — PRÁCTICA BUSCARV'
    ws2['A1'].font = Font(bold=True, size=12, color='7C6FF0')
    ws2.merge_cells('A1:G1')

    h2 = ['NRO_FACTURA','FECHA','COD_PRODUCTO','NOMBRE_PRODUCTO','CANT','PRECIO_UNIT','TOTAL']
    for i, h in enumerate(h2, 1):
        ws2.cell(row=2, column=i).value = h
    style_header(ws2, 2, 7, '7C6FF0')

    facturas = [
        ('F-2025-001','2025-01-05','P001','','3',0,0),
        ('F-2025-002','2025-01-08','P004','','1',0,0),
        ('F-2025-003','2025-01-10','P002','','5',0,0),
        ('F-2025-004','2025-01-15','P007','','2',0,0),
        ('F-2025-005','2025-01-18','P009','','4',0,0),
        ('F-2025-006','2025-01-22','P005','','2',0,0),
        ('F-2025-007','2025-01-25','P012','','1',0,0),
        ('F-2025-008','2025-02-01','P003','','6',0,0),
        ('F-2025-009','2025-02-05','P008','','1',0,0),
        ('F-2025-010','2025-02-10','P010','','10',0,0),
    ]
    for r, row in enumerate(facturas, 3):
        for c, val in enumerate(row, 1):
            ws2.cell(row=r, column=c).value = val
        # Fórmulas para que el estudiante complete
        ws2.cell(row=r, column=4).value = f'=IFERROR(BUSCARV(C{r},Catalogo_Productos!$A:$B,2,0),"No encontrado")'
        ws2.cell(row=r, column=5).value = random.randint(1, 8)
        ws2.cell(row=r, column=6).value = f'=IFERROR(BUSCARV(C{r},Catalogo_Productos!$A:$D,4,0),0)'
        ws2.cell(row=r, column=7).value = f'=E{r}*F{r}'
        style_data_row(ws2, r, 7, alts[r % 2])

    for c in [6, 7]:
        for r in range(3, 13):
            ws2.cell(row=r, column=c).number_format = '"$"#,##0'

    set_col_widths(ws2, [15, 13, 14, 28, 8, 14, 14])
    freeze(ws2)

    # Fila de totales
    ws2.cell(row=14, column=6).value = 'TOTAL VENTAS:'
    ws2.cell(row=14, column=6).font = Font(bold=True)
    ws2.cell(row=14, column=7).value = '=SUM(G3:G13)'
    ws2.cell(row=14, column=7).number_format = '"$"#,##0'
    ws2.cell(row=14, column=7).font = Font(bold=True, color='1AB87A')

    # --- Hoja 3: Guía Actividad ---
    ws3 = wb.create_sheet('GUIA_ACTIVIDAD')
    instrucciones = [
        ('ACTIVIDAD PRÁCTICA — DÍA 1: BUSCARV y BUSCARH', True, 14, '1AB87A'),
        ('', False, 10, None),
        ('OBJETIVO:', True, 11, '333333'),
        ('Usar BUSCARV para traer automáticamente datos del catálogo a la hoja de facturas.', False, 10, None),
        ('', False, 10, None),
        ('PASOS:', True, 11, '333333'),
        ('1. Ve a la hoja "Facturas_Ventas"', False, 10, None),
        ('2. En la columna D (Nombre_Producto), la fórmula ya está puesta como ejemplo en D3.', False, 10, None),
        ('   Revísala: =IFERROR(BUSCARV(C3,Catalogo_Productos!$A:$B,2,0),"No encontrado")', False, 10, '0070C0'),
        ('3. En la columna F (Precio_Unit), la fórmula busca en la columna 4 del catálogo.', False, 10, None),
        ('4. Verifica que los totales (columna G) calculen correctamente.', False, 10, None),
        ('', False, 10, None),
        ('RETO ADICIONAL (si terminas rápido):', True, 11, 'C00000'),
        ('- Crea una hoja "BUSCARH_Prueba" con la tabla de precios en horizontal (por fila)', False, 10, None),
        ('- Usa BUSCARH para buscar el precio de un producto por su código', False, 10, None),
        ('- ¿Cuándo conviene usar BUSCARH vs BUSCARV?', False, 10, None),
        ('', False, 10, None),
        ('ENTREGABLE:', True, 11, '333333'),
        ('Hoja "Facturas_Ventas" con todas las fórmulas funcionando y el total calculado.', False, 10, None),
        ('Guarda como: Actividad_D1_TuNombre.xlsx', False, 10, '7030A0'),
    ]
    for r, (text, bold, size, color) in enumerate(instrucciones, 1):
        cell = ws3.cell(row=r, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color=color or '000000')
    ws3.column_dimensions['A'].width = 80

    save(wb, f'{BASE}/dia01/datos_dia01_BUSCARV.xlsx')

# ── Día 02: INDICE+COINCIDIR / SUMAR.SI ─────────────────────────────────────
def dia02():
    wb = openpyxl.Workbook()

    # Hoja 1: Base empleados
    ws = wb.active
    ws.title = 'Empleados'
    ws['A1'] = 'NÓMINA ADMIN CARIBE S.A.S. — PRÁCTICA INDICE+COINCIDIR'
    ws['A1'].font = Font(bold=True, size=12, color='7C6FF0')
    ws.merge_cells('A1:H1')

    headers = ['ID_EMP','NOMBRE','DEPARTAMENTO','CARGO','CIUDAD','SALARIO','ANTIGUEDAD_AÑOS','BONO']
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i).value = h
    style_header(ws, 2, 8, '7C6FF0')

    empleados = [
        ('E001','Carlos Martínez','Ventas','Vendedor Senior','Barranquilla',2800000,5,280000),
        ('E002','María Rodríguez','Contabilidad','Auxiliar Contable','Barranquilla',1900000,3,0),
        ('E003','Juan García','Logística','Coordinador','Bogotá',3200000,7,320000),
        ('E004','Ana Torres','RRHH','Asistente RRHH','Barranquilla',1800000,2,0),
        ('E005','Pedro Herrera','Ventas','Vendedor Junior','Cartagena',1700000,1,0),
        ('E006','Laura Díaz','Gerencia','Asistente Gerencia','Barranquilla',2500000,4,250000),
        ('E007','Miguel López','Logística','Auxiliar Logístico','Barranquilla',1650000,2,0),
        ('E008','Sandra Castro','Ventas','Vendedora Senior','Bogotá',2900000,6,290000),
        ('E009','Luis Vargas','Contabilidad','Contador Jr','Medellín',2200000,3,0),
        ('E010','Diana Moreno','Ventas','Vendedora Junior','Barranquilla',1700000,1,0),
        ('E011','Roberto Silva','IT','Soporte Técnico','Bogotá',2400000,3,0),
        ('E012','Valeria Gómez','Gerencia','Gerente Comercial','Barranquilla',5500000,8,550000),
    ]

    alts = ['F0E6FF','FFFFFF']
    for r, row in enumerate(empleados, 3):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in [6, 8]:
                cell.number_format = '"$"#,##0'
        style_data_row(ws, r, 8, alts[r % 2])

    set_col_widths(ws, [8, 22, 16, 22, 14, 14, 18, 12])
    freeze(ws)
    autofilter(ws, 'A2:H2')

    # Hoja 2: Tabla Ventas para búsqueda bidireccional
    ws2 = wb.create_sheet('Ventas_Trimestre')
    ws2['A1'] = 'VENTAS POR CIUDAD Y MES — BÚSQUEDA BIDIRECCIONAL'
    ws2['A1'].font = Font(bold=True, size=12, color='7C6FF0')
    ws2.merge_cells('A1:E1')

    ws2.cell(row=2, column=1).value = 'CIUDAD / MES'
    meses = ['Enero', 'Febrero', 'Marzo']
    ciudades = ['Barranquilla', 'Bogotá', 'Cartagena', 'Medellín']
    datos_ventas = [
        [38500000, 41200000, 43800000],
        [54000000, 58300000, 61200000],
        [22100000, 24500000, 26800000],
        [28000000, 30500000, 33200000],
    ]
    style_header(ws2, 2, 4, '7C6FF0')
    for i, m in enumerate(meses, 2):
        ws2.cell(row=2, column=i).value = m
    for r, (ciudad, ventas) in enumerate(zip(ciudades, datos_ventas), 3):
        ws2.cell(row=r, column=1).value = ciudad
        for c, v in enumerate(ventas, 2):
            cell = ws2.cell(row=r, column=c, value=v)
            cell.number_format = '"$"#,##0'
        style_data_row(ws2, r, 4, alts[r % 2])
    set_col_widths(ws2, [18, 16, 16, 16])

    # Hoja 3: Panel búsqueda
    ws3 = wb.create_sheet('Panel_Busqueda')
    ws3['A1'] = 'PANEL DE BÚSQUEDA — PRACTICA INDICE+COINCIDIR'
    ws3['A1'].font = Font(bold=True, size=12, color='7C6FF0')
    ws3.merge_cells('A1:D1')
    ws3['A3'] = 'Ciudad a buscar:'
    ws3['A4'] = 'Mes a buscar:'
    ws3['B3'] = 'Barranquilla'
    ws3['B4'] = 'Febrero'
    ws3['B3'].fill = PatternFill('solid', fgColor='FFF2CC')
    ws3['B4'].fill = PatternFill('solid', fgColor='FFF2CC')
    ws3['A6'] = 'RESULTADO — Ventas:'
    ws3['A6'].font = Font(bold=True)
    ws3['B6'] = "=IFERROR(INDICE(Ventas_Trimestre!B3:D6,COINCIDIR(B3,Ventas_Trimestre!A3:A6,0),COINCIDIR(B4,Ventas_Trimestre!B2:D2,0)),\"No encontrado\")"
    ws3['B6'].font = Font(bold=True, color='7C6FF0', size=14)
    ws3['B6'].number_format = '"$"#,##0'
    ws3['A8'] = 'Instrucción: Cambia B3 y B4 para ver diferentes resultados'
    ws3['A8'].font = Font(italic=True, color='888888')
    set_col_widths(ws3, [22, 22, 16, 16])

    # Hoja 4: SUMAR.SI práctica
    ws4 = wb.create_sheet('Practica_SUMАРSI')
    ws4.title = 'Practica_SUMARSI'
    ws4['A1'] = 'ANÁLISIS DE NÓMINA CON SUMAR.SI — CONTAR.SI — PROMEDIO.SI'
    ws4['A1'].font = Font(bold=True, size=12, color='7C6FF0')
    ws4.merge_cells('A1:D1')

    ws4['A3'] = 'RESUMEN POR DEPARTAMENTO'
    ws4['A3'].font = Font(bold=True, size=11)
    headers4 = ['DEPARTAMENTO', 'TOTAL_NOMINA', 'NRO_EMPLEADOS', 'SALARIO_PROMEDIO']
    for i, h in enumerate(headers4, 1):
        ws4.cell(row=4, column=i).value = h
    style_header(ws4, 4, 4, '7C6FF0')

    deptos = ['Ventas', 'Contabilidad', 'Logística', 'RRHH', 'Gerencia', 'IT']
    for r, dep in enumerate(deptos, 5):
        ws4.cell(row=r, column=1).value = dep
        ws4.cell(row=r, column=2).value = f'=SUMAR.SI(Empleados!$C:$C,A{r},Empleados!$F:$F)'
        ws4.cell(row=r, column=2).number_format = '"$"#,##0'
        ws4.cell(row=r, column=3).value = f'=CONTAR.SI(Empleados!$C:$C,A{r})'
        ws4.cell(row=r, column=4).value = f'=IFERROR(B{r}/C{r},0)'
        ws4.cell(row=r, column=4).number_format = '"$"#,##0'
        style_data_row(ws4, r, 4, alts[r % 2])

    set_col_widths(ws4, [18, 18, 18, 20])

    # Guía actividad
    ws5 = wb.create_sheet('GUIA_ACTIVIDAD')
    guia = [
        ('ACTIVIDAD PRÁCTICA — DÍA 2: INDICE+COINCIDIR y SUMAR.SI', True, 14, '7C6FF0'),
        ('', False, 10, None),
        ('PARTE 1 — BÚSQUEDA BIDIRECCIONAL (20 min)', True, 11, '333333'),
        ('1. Ve a la hoja "Panel_Busqueda"', False, 10, None),
        ('2. Cambia la ciudad en B3 a "Bogotá" — ¿qué pasa con el resultado?', False, 10, None),
        ('3. Cambia el mes en B4 a "Marzo"', False, 10, None),
        ('4. Explica con tus palabras cómo funciona la fórmula INDICE+COINCIDIR', False, 10, None),
        ('', False, 10, None),
        ('PARTE 2 — ANÁLISIS DE NÓMINA (20 min)', True, 11, '333333'),
        ('1. Ve a la hoja "Practica_SUMARSI"', False, 10, None),
        ('2. Verifica que las fórmulas SUMAR.SI y CONTAR.SI funcionen correctamente', False, 10, None),
        ('3. Agrega una fila "TOTAL EMPRESA" al final con la suma de toda la nómina', False, 10, None),
        ('4. Crea una columna adicional "% DEL TOTAL" que muestre el porcentaje de cada depto', False, 10, None),
        ('', False, 10, None),
        ('RETO: SUMAR.SI.CONJUNTO (15 min)', True, 11, 'C00000'),
        ('En una celda aparte, calcula el total de salarios de vendedores en Barranquilla', False, 10, None),
        ('Pista: =SUMAR.SI.CONJUNTO(Empleados!F:F, Empleados!C:C,"Ventas", Empleados!E:E,"Barranquilla")', False, 10, '0070C0'),
        ('', False, 10, None),
        ('Guarda como: Actividad_D2_TuNombre.xlsx', False, 10, '7030A0'),
    ]
    for r, (text, bold, size, color) in enumerate(guia, 1):
        cell = ws5.cell(row=r, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color=color or '000000')
    ws5.column_dimensions['A'].width = 85

    save(wb, f'{BASE}/dia02/datos_dia02_INDICE_COINCIDIR.xlsx')

# ── Día 03: Funciones Financieras y Texto ────────────────────────────────────
def dia03():
    wb = openpyxl.Workbook()

    # Hoja 1: Funciones Financieras
    ws = wb.active
    ws.title = 'Funciones_Financieras'
    ws['A1'] = 'CALCULADORA FINANCIERA — PAGO, VA, VF, TASA, NPER'
    ws['A1'].font = Font(bold=True, size=12, color='F0B429')
    ws.merge_cells('A1:D1')

    # Bloque PAGO
    ws['A3'] = 'FUNCIÓN PAGO — ¿Cuánto pago por mes?'
    ws['A3'].font = Font(bold=True, size=11, color='F0B429')
    datos_pago = [
        ('Monto del crédito (PV):', 50000000, '"$"#,##0'),
        ('Tasa mensual (RATE):', 0.012, '0.00%'),
        ('Número de cuotas (NPER):', 36, '0'),
        ('CUOTA MENSUAL (PAGO):', '=PAGO(B6,B7,B5)*-1', '"$"#,##0'),
        ('TOTAL A PAGAR:', '=B8*B7', '"$"#,##0'),
        ('TOTAL INTERESES:', '=B9-B5', '"$"#,##0'),
    ]
    labels_pago = ['Monto del crédito:', 'Tasa mensual:', 'Nro cuotas:', 'CUOTA MENSUAL:', 'TOTAL A PAGAR:', 'TOTAL INTERESES:']
    for i, (label, val, fmt) in enumerate(datos_pago):
        r = i + 5
        ws.cell(row=r, column=1).value = label
        cell = ws.cell(row=r, column=2, value=val)
        cell.number_format = fmt
        if label.startswith('CUOTA') or label.startswith('TOTAL'):
            ws.cell(row=r, column=1).font = Font(bold=True)
            cell.font = Font(bold=True, color='F0B429')
            cell.fill = PatternFill('solid', fgColor='1a1500')

    # Bloque VF
    ws['A12'] = 'FUNCIÓN VF — ¿Cuánto tendré si ahorro?'
    ws['A12'].font = Font(bold=True, size=11, color='F0B429')
    ws['A13'] = 'Ahorro mensual (PMT):'
    ws['B13'] = 500000
    ws['B13'].number_format = '"$"#,##0'
    ws['A14'] = 'Tasa mensual (RATE):'
    ws['B14'] = 0.008
    ws['B14'].number_format = '0.00%'
    ws['A15'] = 'Número de meses (NPER):'
    ws['B15'] = 24
    ws['A16'] = 'VALOR FUTURO (VF):'
    ws['B16'] = '=VF(B14,B15,-B13,0,0)'
    ws['B16'].number_format = '"$"#,##0'
    ws['B16'].font = Font(bold=True, color='1AB87A')

    # Bloque VA
    ws['A18'] = 'FUNCIÓN VA — ¿Cuánto vale hoy una renta futura?'
    ws['A18'].font = Font(bold=True, size=11, color='F0B429')
    ws['A19'] = 'Renta mensual (PMT):'
    ws['B19'] = 800000
    ws['B19'].number_format = '"$"#,##0'
    ws['A20'] = 'Tasa de descuento mensual:'
    ws['B20'] = 0.01
    ws['B20'].number_format = '0.00%'
    ws['A21'] = 'Número de períodos:'
    ws['B21'] = 12
    ws['A22'] = 'VALOR ACTUAL (VA):'
    ws['B22'] = '=VA(B20,B21,-B19,0,0)'
    ws['B22'].number_format = '"$"#,##0'
    ws['B22'].font = Font(bold=True, color='1AB87A')

    set_col_widths(ws, [32, 18, 16, 16])

    # Hoja 2: Tabla amortización
    ws2 = wb.create_sheet('Tabla_Amortizacion')
    ws2['A1'] = 'TABLA DE AMORTIZACIÓN — CRÉDITO $50,000,000 — 12 MESES — 1.2% MES'
    ws2['A1'].font = Font(bold=True, size=11, color='F0B429')
    ws2.merge_cells('A1:F1')
    headers = ['CUOTA','SALDO_INICIAL','INTERÉS','ABONO_CAPITAL','PAGO_TOTAL','SALDO_FINAL']
    for i, h in enumerate(headers, 1):
        ws2.cell(row=2, column=i).value = h
    style_header(ws2, 2, 6, 'F0B429', '1a1500')

    tasa, n, pv = 0.012, 12, 50000000
    import math
    cuota = pv * tasa / (1 - (1 + tasa) ** -n)
    saldo = pv
    alts = ['FFF8E1', 'FFFFFF']
    for r in range(n):
        interes = saldo * tasa
        capital = cuota - interes
        saldo_fin = saldo - capital
        row_data = [r+1, saldo, interes, capital, cuota, max(saldo_fin, 0)]
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r+3, column=c, value=round(val, 0))
            if c > 1:
                cell.number_format = '"$"#,##0'
        style_data_row(ws2, r+3, 6, alts[r % 2])
        saldo = max(saldo_fin, 0)

    # Totales
    ws2.cell(row=n+4, column=1).value = 'TOTALES'
    ws2.cell(row=n+4, column=1).font = Font(bold=True)
    for c in [3, 4, 5]:
        cell = ws2.cell(row=n+4, column=c)
        col = get_column_letter(c)
        cell.value = f'=SUM({col}3:{col}{n+2})'
        cell.number_format = '"$"#,##0'
        cell.font = Font(bold=True, color='F0B429')

    set_col_widths(ws2, [8, 18, 16, 18, 16, 18])
    freeze(ws2)

    # Hoja 3: Funciones de Texto
    ws3 = wb.create_sheet('Funciones_Texto')
    ws3['A1'] = 'FUNCIONES DE TEXTO — CASOS PRÁCTICOS'
    ws3['A1'].font = Font(bold=True, size=12, color='F0B429')
    ws3.merge_cells('A1:C1')

    datos_texto = [
        ('DATO ORIGINAL', 'FUNCIÓN', 'RESULTADO'),
        ('  juan garcia  ', '=NOMPROPIO(ESPACIOS(A4))', ''),
        ('juan.garcia@empresa.com', '=IZQUIERDA(A5,ENCONTRAR("@",A5)-1)', ''),
        ('CARLOS MARTINEZ PEREZ', '=EXTRAE(A6,8,8)', ''),
        ('Factura N° 2025-001', '=SUSTITUIR(A7,"°","No.")', ''),
        ('Barranquilla,Colombia,Atlántico', '=IZQUIERDA(A8,ENCONTRAR(",",A8)-1)', ''),
        ('3001234567', '=CONCATENAR("(",IZQUIERDA(A9,3),") ",EXTRAE(A9,4,3),"-",DERECHA(A9,4))', ''),
        ('precio: 125000 pesos', '=VALOR(EXTRAE(A10,9,6))', ''),
    ]

    for r, row in enumerate(datos_texto, 3):
        for c, val in enumerate(row, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            if r == 3:
                cell.font = Font(bold=True, color='F0B429')
                cell.fill = PatternFill('solid', fgColor='2a1500')
            elif c == 2:
                cell.font = Font(color='0070C0', size=9)

    set_col_widths(ws3, [35, 55, 25])

    # Guía
    ws4 = wb.create_sheet('GUIA_ACTIVIDAD')
    guia = [
        ('ACTIVIDAD PRÁCTICA — DÍA 3: Funciones Financieras, Texto y Matriciales', True, 13, 'F0B429'),
        ('', False, 10, None),
        ('PARTE 1 — SIMULACIÓN DE CRÉDITO (20 min)', True, 11, '333333'),
        ('1. Ve a la hoja "Funciones_Financieras"', False, 10, None),
        ('2. Cambia el monto a $80,000,000 y 48 cuotas — ¿cuánto sería la cuota?', False, 10, None),
        ('3. ¿A qué tasa quedaría una cuota de $2,500,000 para ese mismo crédito? (usa TASA)', False, 10, None),
        ('4. Crea tu propio escenario: ¿cuánto necesitas ahorrar por mes para tener $20M en 3 años?', False, 10, None),
        ('', False, 10, None),
        ('PARTE 2 — TEXTO Y DATOS (15 min)', True, 11, '333333'),
        ('1. En la hoja "Funciones_Texto", revisa cada fórmula en la columna B', False, 10, None),
        ('2. En la columna C escribe el resultado esperado ANTES de ver la fórmula', False, 10, None),
        ('3. Crea 3 filas adicionales con tus propios ejemplos de datos "sucios"', False, 10, None),
        ('', False, 10, None),
        ('PARTE 3 — TABLA DE AMORTIZACIÓN (25 min)', True, 11, '333333'),
        ('1. Ve a la hoja "Tabla_Amortizacion"', False, 10, None),
        ('2. Verifica que la suma de ABONO_CAPITAL sea igual al monto original', False, 10, None),
        ('3. Crea una tabla idéntica pero para 24 meses en una nueva hoja', False, 10, None),
        ('4. ¿Cuánto más en intereses se paga en 24 meses vs 12 meses?', False, 10, None),
        ('', False, 10, None),
        ('Guarda como: Actividad_D3_TuNombre.xlsx', False, 10, '7030A0'),
    ]
    for r, (text, bold, size, color) in enumerate(guia, 1):
        ws4.cell(row=r, column=1).value = text
        ws4.cell(row=r, column=1).font = Font(bold=bold, size=size, color=color or '000000')
    ws4.column_dimensions['A'].width = 85

    save(wb, f'{BASE}/dia03/datos_dia03_Financiero_Texto.xlsx')

# ── Día 04: Tablas Dinámicas ──────────────────────────────────────────────────
def dia04():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Datos_Ventas'
    ws['A1'] = 'BASE DE DATOS VENTAS — ADMIN CARIBE S.A.S. — 2025'
    ws['A1'].font = Font(bold=True, size=12, color='00D4AA')
    ws.merge_cells('A1:I1')

    headers = ['ID_VENTA','FECHA','MES','VENDEDOR','CIUDAD','PRODUCTO','CATEGORIA','CANTIDAD','MONTO']
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i).value = h
    style_header(ws, 2, 9, '00D4AA', '002a20')

    vendedores = ['Carlos Martínez','Sandra Castro','Pedro Herrera','Diana Moreno','Luis Vargas']
    ciudades = ['Barranquilla','Bogotá','Cartagena','Medellín','Barranquilla']
    productos_cat = [
        ('Resma Papel','Papelería'),('Tóner HP','Tecnología'),('Carpeta AZ','Archivos'),
        ('USB 32GB','Tecnología'),('Calculadora','Oficina'),('Grapadora','Herramientas'),
        ('Bolígrafo x12','Papelería'),('Folder Manila','Papelería'),
    ]
    meses_num = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
                 7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',11:'Noviembre',12:'Diciembre'}

    random.seed(42)
    alts = ['E6FDF8','FFFFFF']
    for row_i in range(100):
        r = row_i + 3
        mes = random.randint(1, 12)
        dia = random.randint(1, 28)
        fecha = f'2025-{mes:02d}-{dia:02d}'
        vend_i = random.randint(0, 4)
        prod_i = random.randint(0, 7)
        prod, cat = productos_cat[prod_i]
        cant = random.randint(1, 20)
        precio = random.randint(5, 200) * 1000
        row_data = [f'V{row_i+1:03d}', fecha, meses_num[mes],
                    vendedores[vend_i], ciudades[vend_i],
                    prod, cat, cant, cant * precio]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 9:
                cell.number_format = '"$"#,##0'
        style_data_row(ws, r, 9, alts[row_i % 2])

    set_col_widths(ws, [10, 13, 12, 22, 14, 20, 14, 11, 16])
    freeze(ws)
    autofilter(ws, 'A2:I2')

    # Hoja instrucciones
    ws2 = wb.create_sheet('GUIA_ACTIVIDAD')
    guia = [
        ('ACTIVIDAD PRÁCTICA — DÍA 4: Tablas Dinámicas', True, 14, '00D4AA'),
        ('', False, 10, None),
        ('PARTE 1 — CREAR TABLAS DINÁMICAS (30 min)', True, 11, '333333'),
        ('1. Selecciona todos los datos de la hoja "Datos_Ventas" (A2:I102)', False, 10, None),
        ('2. Insertar → Tabla dinámica → Nueva hoja', False, 10, None),
        ('3. Tabla dinámica 1: Ventas por CIUDAD (filas) y MES (columnas) — valores: SUMA de MONTO', False, 10, None),
        ('4. Tabla dinámica 2: Ventas por VENDEDOR — valores: SUMA MONTO y CONTAR CANTIDAD', False, 10, None),
        ('5. Tabla dinámica 3: Top de CATEGORÍAS por monto total', False, 10, None),
        ('', False, 10, None),
        ('PARTE 2 — FILTROS Y SLICERS (15 min)', True, 11, '333333'),
        ('1. En la Tabla dinámica 1, inserta un Slicer de "CIUDAD"', False, 10, None),
        ('2. Conecta el mismo slicer a la Tabla dinámica 2', False, 10, None),
        ('3. Inserta una Línea de Tiempo para el campo "FECHA"', False, 10, None),
        ('4. Filtra para ver solo el primer trimestre (Enero-Marzo)', False, 10, None),
        ('', False, 10, None),
        ('PARTE 3 — CONFIGURACIÓN AVANZADA (15 min)', True, 11, '333333'),
        ('1. En la tabla de vendedores, cambia el cálculo de MONTO a "% del total general"', False, 10, None),
        ('2. Agrega un campo calculado: "Comisión = MONTO * 3%"', False, 10, None),
        ('3. Aplica un diseño de tabla dinámica que se vea profesional', False, 10, None),
        ('', False, 10, None),
        ('Guarda como: Actividad_D4_TuNombre.xlsx', False, 10, '7030A0'),
    ]
    for r, (text, bold, size, color) in enumerate(guia, 1):
        ws2.cell(row=r, column=1).value = text
        ws2.cell(row=r, column=1).font = Font(bold=bold, size=size, color=color or '000000')
    ws2.column_dimensions['A'].width = 85

    save(wb, f'{BASE}/dia04/datos_dia04_Tablas_Dinamicas.xlsx')

# ── Día 05: Gráficos y Power Query ──────────────────────────────────────────
def dia05():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Ventas_para_Graficos'

    ws['A1'] = 'DATOS PARA GRÁFICOS DINÁMICOS — ADMIN CARIBE'
    ws['A1'].font = Font(bold=True, size=12, color='C084FC')
    ws.merge_cells('A1:E1')

    # Resumen mensual por ciudad (como si viniera de tabla dinámica)
    meses = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
    ws.cell(row=2, column=1).value = 'MES'
    ciudades_g = ['Barranquilla','Bogotá','Cartagena']
    for i, c in enumerate(ciudades_g, 2):
        ws.cell(row=2, column=i).value = c
    style_header(ws, 2, 4, 'C084FC', '1a0030')

    random.seed(10)
    alts = ['F5E6FF','FFFFFF']
    base_vals = [38000000, 54000000, 22000000]
    for r, mes in enumerate(meses, 3):
        ws.cell(row=r, column=1).value = mes
        for c_i, base in enumerate(base_vals, 2):
            v = base + random.randint(-5000000, 8000000)
            cell = ws.cell(row=r, column=c_i, value=v)
            cell.number_format = '"$"#,##0'
        style_data_row(ws, r, 4, alts[r % 2])

    set_col_widths(ws, [8, 18, 16, 16])

    # Hoja datos sucios para Power Query
    ws2 = wb.create_sheet('Datos_Sucios_PQ')
    ws2['A1'] = ' REPORTE VENTAS '  # espacios a propósito
    ws2['B1'] = 'PERIODO'
    ws2['C1'] = 'CIUDAD'
    ws2['D1'] = 'TOTAL'
    style_header(ws2, 1, 4, 'C084FC', '1a0030')

    datos_sucios = [
        ('Carlos MARTINEZ', 'Enero-2025', ' Barranquilla ', '38500000'),
        ('SANDRA Castro', 'Febrero-2025', ' BOGOTÁ ', '41200000'),
        ('Pedro herrera', 'Enero-2025', 'cartagena', '22100000'),
        ('diana Moreno', 'Marzo-2025', 'BARRANQUILLA', '43800000'),
        ('LUIS vargas', 'Febrero-2025', 'Medellín ', '30500000'),
        ('carlos martinez', 'Marzo-2025', 'Barranquilla', '38900000'),
        ('Sandra CASTRO', 'Enero-2025', 'bogotá', '54000000'),
    ]
    for r, row in enumerate(datos_sucios, 2):
        for c, val in enumerate(row, 1):
            ws2.cell(row=r, column=c, value=val)
        style_data_row(ws2, r, 4, alts[r % 2])

    ws2['E1'] = '← Nota: estos datos tienen espacios, mayúsculas inconsistentes. Power Query los limpia.'
    ws2['E1'].font = Font(italic=True, color='888888', size=9)
    set_col_widths(ws2, [22, 14, 16, 14, 60])

    ws3 = wb.create_sheet('GUIA_ACTIVIDAD')
    guia = [
        ('ACTIVIDAD PRÁCTICA — DÍA 5: Gráficos Dinámicos y Power Query', True, 13, 'C084FC'),
        ('', False, 10, None),
        ('PARTE 1 — GRÁFICOS (25 min)', True, 11, '333333'),
        ('1. Selecciona A2:D14 en la hoja "Ventas_para_Graficos"', False, 10, None),
        ('2. Crea un gráfico de columnas agrupadas por ciudad por mes', False, 10, None),
        ('3. Crea un gráfico de líneas para ver la tendencia de Barranquilla en el año', False, 10, None),
        ('4. Dale título: "Evolución de Ventas 2025 — ADMIN CARIBE"', False, 10, None),
        ('5. Aplica colores corporativos y etiquetas de datos', False, 10, None),
        ('', False, 10, None),
        ('PARTE 2 — POWER QUERY (30 min)', True, 11, '333333'),
        ('1. Ve a la hoja "Datos_Sucios_PQ" — observa los problemas de calidad', False, 10, None),
        ('2. Selecciona los datos → Datos → Obtener y transformar → Desde tabla', False, 10, None),
        ('3. En Power Query: quita espacios del nombre (Transformar → Recortar)', False, 10, None),
        ('4. Estandariza la columna Ciudad: Transformar → Formato → Escribir en mayúscula/minúscula', False, 10, None),
        ('5. Cambia el tipo de columna TOTAL a Número entero', False, 10, None),
        ('6. Cerrar y cargar → verifica que los datos están limpios en Excel', False, 10, None),
        ('', False, 10, None),
        ('Guarda como: Actividad_D5_TuNombre.xlsx', False, 10, '7030A0'),
    ]
    for r, (text, bold, size, color) in enumerate(guia, 1):
        ws3.cell(row=r, column=1).value = text
        ws3.cell(row=r, column=1).font = Font(bold=bold, size=size, color=color or '000000')
    ws3.column_dimensions['A'].width = 85

    save(wb, f'{BASE}/dia05/datos_dia05_Graficos_PowerQuery.xlsx')

# ── Día 06: Macros y Dashboard ───────────────────────────────────────────────
def dia06():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Datos_Dashboard'

    ws['A1'] = 'DATOS BASE PARA DASHBOARD — ADMIN CARIBE S.A.S.'
    ws['A1'].font = Font(bold=True, size=12, color='4ADE80')
    ws.merge_cells('A1:H1')

    headers = ['VENDEDOR','ENERO','FEBRERO','MARZO','TOTAL_Q1','META_Q1','% CUMPLIMIENTO','ESTADO']
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i).value = h
    style_header(ws, 2, 8, '4ADE80', '002000')

    vendedores_data = [
        ('Carlos Martínez',38500000,41200000,43800000,5000000),
        ('Sandra Castro',54000000,58300000,61200000,5000000),
        ('Pedro Herrera',22100000,24500000,26800000,4000000),
        ('Diana Moreno',18000000,19500000,21000000,3500000),
        ('Luis Vargas',12000000,13500000,14800000,2500000),
    ]
    metas = [120000000, 180000000, 75000000, 60000000, 42000000]

    alts = ['E6FFEE','FFFFFF']
    for r_i, (row, meta) in enumerate(zip(vendedores_data, metas)):
        r = r_i + 3
        nombre, e, f, m, _ = row
        total = e + f + m
        pct = total / meta
        estado = 'CUMPLIDO' if pct >= 1 else ('EN RIESGO' if pct >= 0.85 else 'BAJO META')
        row_data = [nombre, e, f, m, total, meta, pct, estado]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in [2, 3, 4, 5, 6]:
                cell.number_format = '"$"#,##0'
            elif c == 7:
                cell.number_format = '0.0%'
            elif c == 8:
                if val == 'CUMPLIDO':
                    cell.fill = PatternFill('solid', fgColor='C6EFCE')
                    cell.font = Font(color='276221', bold=True)
                elif val == 'EN RIESGO':
                    cell.fill = PatternFill('solid', fgColor='FFEB9C')
                    cell.font = Font(color='9C6500', bold=True)
                else:
                    cell.fill = PatternFill('solid', fgColor='FFC7CE')
                    cell.font = Font(color='9C0006', bold=True)
        style_data_row(ws, r, 7, alts[r_i % 2])

    # Totales
    ws.cell(row=9, column=1).value = 'TOTAL EMPRESA'
    ws.cell(row=9, column=1).font = Font(bold=True)
    for c in [2, 3, 4, 5, 6]:
        col = get_column_letter(c)
        cell = ws.cell(row=9, column=c)
        cell.value = f'=SUM({col}3:{col}8)'
        cell.number_format = '"$"#,##0'
        cell.font = Font(bold=True, color='4ADE80')

    set_col_widths(ws, [22, 15, 15, 15, 16, 14, 18, 14])

    # Hoja KPIs
    ws2 = wb.create_sheet('KPIs_Dashboard')
    ws2['A1'] = 'PANEL KPIs — DASHBOARD EJECUTIVO'
    ws2['A1'].font = Font(bold=True, size=14, color='4ADE80')
    ws2.merge_cells('A1:D1')
    ws2.row_dimensions[1].height = 30

    kpis = [
        ('VENTAS TOTALES Q1', '=SUM(Datos_Dashboard!E3:E8)', '"$"#,##0'),
        ('MEJOR VENDEDOR', '=INDICE(Datos_Dashboard!A3:A8,COINCIDIR(MAX(Datos_Dashboard!E3:E8),Datos_Dashboard!E3:E8,0))', '@'),
        ('% CUMPLIMIENTO PROMEDIO', '=PROMEDIO(Datos_Dashboard!G3:G8)', '0.0%'),
        ('VENDEDORES EN META', '=CONTAR.SI(Datos_Dashboard!H3:H8,"CUMPLIDO")', '0'),
    ]
    for r, (label, formula, fmt) in enumerate(kpis, 3):
        ws2.cell(row=r, column=1).value = label
        ws2.cell(row=r, column=1).font = Font(bold=True, size=10)
        cell = ws2.cell(row=r, column=2, value=formula)
        cell.number_format = fmt
        cell.font = Font(bold=True, size=14, color='4ADE80')
        ws2.row_dimensions[r].height = 35

    set_col_widths(ws2, [32, 28, 16, 16])

    ws3 = wb.create_sheet('GUIA_ACTIVIDAD')
    guia = [
        ('ACTIVIDAD PRÁCTICA — DÍA 6: Macros y Dashboard Ejecutivo', True, 14, '4ADE80'),
        ('', False, 10, None),
        ('PARTE 1 — MACRO DE FORMATO (20 min)', True, 11, '333333'),
        ('1. Abre el Editor VBA: Alt+F11', False, 10, None),
        ('2. Inserta un módulo: Insertar → Módulo', False, 10, None),
        ('3. Escribe (o copia) la siguiente macro:', False, 10, None),
        ('   Sub FormatoReporte()', False, 10, '0070C0'),
        ('     Rows("1:1").Font.Bold = True', False, 10, '0070C0'),
        ('     Rows("1:1").Interior.Color = RGB(26, 184, 122)', False, 10, '0070C0'),
        ('     Cells.EntireColumn.AutoFit', False, 10, '0070C0'),
        ('     MsgBox "Formato aplicado correctamente!"', False, 10, '0070C0'),
        ('   End Sub', False, 10, '0070C0'),
        ('4. Ejecuta con F5 — verifica que funciona', False, 10, None),
        ('5. Asigna la macro a un botón en la hoja "KPIs_Dashboard"', False, 10, None),
        ('', False, 10, None),
        ('PARTE 2 — DASHBOARD (40 min)', True, 11, '333333'),
        ('1. En la hoja "KPIs_Dashboard", construye el panel visual con los 4 KPIs', False, 10, None),
        ('2. Agrega un gráfico de barras de los vendedores (fuente: Datos_Dashboard)', False, 10, None),
        ('3. Agrega formato condicional: si % cumplimiento < 85% → celda roja', False, 10, None),
        ('4. Protege la hoja "Datos_Dashboard" con contraseña: "Admin2025"', False, 10, None),
        ('5. Guarda como .xlsm (libro habilitado para macros)', False, 10, None),
        ('', False, 10, None),
        ('Guarda como: Dashboard_AdminCaribe_TuNombre.xlsm', False, 10, '7030A0'),
    ]
    for r, (text, bold, size, color) in enumerate(guia, 1):
        ws3.cell(row=r, column=1).value = text
        ws3.cell(row=r, column=1).font = Font(bold=bold, size=size, color=color or '000000')
    ws3.column_dimensions['A'].width = 85

    save(wb, f'{BASE}/dia06/datos_dia06_Macros_Dashboard.xlsx')

# ── Día 09: Access / Base de Datos ───────────────────────────────────────────
def dia09():
    wb = openpyxl.Workbook()

    # Hoja clientes para importar a Access
    ws = wb.active
    ws.title = 'tbl_Clientes'
    ws['A1'] = 'TABLA CLIENTES — IMPORTAR A ACCESS'
    ws['A1'].font = Font(bold=True, size=12, color='E879F9')
    ws.merge_cells('A1:F1')

    headers = ['ID_CLIENTE','NOMBRE_CLIENTE','CIUDAD','TELEFONO','EMAIL','FECHA_REGISTRO']
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i).value = h
    style_header(ws, 2, 6, 'E879F9', '1a0020')

    clientes = [
        (1,'Distribuciones Caribe S.A.','Barranquilla','(605) 385-1234','info@discaribe.com','2023-01-15'),
        (2,'Logística del Norte Ltda.','Barranquilla','(605) 372-5678','ventas@lognorte.co','2023-02-20'),
        (3,'Comercial Pacífico S.A.S.','Bogotá','(601) 234-9012','contacto@compaci.co','2023-03-10'),
        (4,'Inversiones Costa S.A.','Cartagena','(605) 664-3456','info@invcosta.com','2023-04-05'),
        (5,'Suministros Atlántico','Barranquilla','(605) 391-7890','sumin@atlantico.co','2023-05-18'),
        (6,'Tech Solutions Colombia','Bogotá','(601) 456-1234','ventas@techcol.co','2023-06-22'),
        (7,'Alimentos del Caribe','Barranquilla','(605) 385-5678','info@alicaribe.com','2023-07-30'),
        (8,'Ferretería Industrial S.A.','Medellín','(604) 567-9012','ferr@industrial.co','2023-08-14'),
    ]
    alts = ['F9E6FF','FFFFFF']
    for r, row in enumerate(clientes, 3):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
        style_data_row(ws, r, 6, alts[r % 2])

    set_col_widths(ws, [12, 30, 16, 18, 30, 16])

    # Hoja productos
    ws2 = wb.create_sheet('tbl_Productos')
    headers2 = ['ID_PRODUCTO','NOMBRE_PRODUCTO','CATEGORIA','PRECIO_UNITARIO','STOCK']
    for i, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=i).value = h
    style_header(ws2, 1, 5, 'E879F9', '1a0020')

    prods = [
        (1,'Resma Papel Carta x500','Papelería',18500,500),
        (2,'Tóner HP LaserJet 85A','Tecnología',185000,50),
        (3,'Carpeta AZ Oficio','Archivos',8500,300),
        (4,'Memoria USB 32GB','Tecnología',35000,120),
        (5,'Calculadora Científica','Oficina',95000,40),
    ]
    for r, row in enumerate(prods, 2):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c == 4:
                cell.number_format = '"$"#,##0'
        style_data_row(ws2, r, 5, alts[r % 2])

    set_col_widths(ws2, [14, 28, 14, 18, 10])

    # Hoja pedidos
    ws3 = wb.create_sheet('tbl_Pedidos')
    headers3 = ['ID_PEDIDO','ID_CLIENTE','ID_PRODUCTO','FECHA_PEDIDO','CANTIDAD','ESTADO']
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=i).value = h
    style_header(ws3, 1, 6, 'E879F9', '1a0020')

    pedidos = [
        (1,1,1,'2025-01-10',10,'Entregado'),
        (2,2,2,'2025-01-15',2,'Entregado'),
        (3,3,4,'2025-01-20',5,'Pendiente'),
        (4,1,3,'2025-02-05',20,'Entregado'),
        (5,4,5,'2025-02-10',3,'Pendiente'),
        (6,2,1,'2025-02-15',15,'En tránsito'),
        (7,5,2,'2025-03-01',1,'Entregado'),
        (8,6,4,'2025-03-10',8,'Pendiente'),
        (9,3,3,'2025-03-15',12,'En tránsito'),
        (10,1,5,'2025-03-20',2,'Entregado'),
    ]
    for r, row in enumerate(pedidos, 2):
        for c, val in enumerate(row, 1):
            ws3.cell(row=r, column=c, value=val)
        style_data_row(ws3, r, 6, alts[r % 2])

    set_col_widths(ws3, [12, 12, 14, 15, 11, 14])

    ws4 = wb.create_sheet('GUIA_ACTIVIDAD')
    guia = [
        ('ACTIVIDAD PRÁCTICA — DÍA 9: Bases de Datos y Access', True, 14, 'E879F9'),
        ('', False, 10, None),
        ('PARTE 1 — IMPORTAR A ACCESS (25 min)', True, 11, '333333'),
        ('1. Abre Microsoft Access → Nuevo → Base de datos en blanco: "AdminCaribe.accdb"', False, 10, None),
        ('2. Datos externos → Excel → Importa la hoja "tbl_Clientes" de este archivo', False, 10, None),
        ('3. Asegúrate de marcar "La primera fila contiene encabezados"', False, 10, None),
        ('4. Define ID_CLIENTE como clave principal', False, 10, None),
        ('5. Repite para "tbl_Productos" y "tbl_Pedidos"', False, 10, None),
        ('', False, 10, None),
        ('PARTE 2 — RELACIONES (20 min)', True, 11, '333333'),
        ('1. Herramientas de BD → Relaciones → Agrega las 3 tablas', False, 10, None),
        ('2. Arrastra ID_CLIENTE de tbl_Clientes hacia ID_CLIENTE en tbl_Pedidos', False, 10, None),
        ('3. Activa "Exigir integridad referencial" → Crear', False, 10, None),
        ('4. Repite con ID_PRODUCTO', False, 10, None),
        ('5. Captura de pantalla del diagrama de relaciones → guardar en tu portafolio', False, 10, None),
        ('', False, 10, None),
        ('PARTE 3 — CONSULTA (15 min)', True, 11, '333333'),
        ('1. Crear → Diseño de consulta', False, 10, None),
        ('2. Agrega tbl_Clientes, tbl_Pedidos, tbl_Productos', False, 10, None),
        ('3. Arrastra: NombreCliente, NombreProducto, Cantidad, FechaPedido, Estado', False, 10, None),
        ('4. En criterios de "Estado" escribe: Pendiente', False, 10, None),
        ('5. Ejecuta con ! → guarda como "qry_Pedidos_Pendientes"', False, 10, None),
        ('', False, 10, None),
        ('Archivo de Access: AdminCaribe_TuNombre.accdb', False, 10, '7030A0'),
    ]
    for r, (text, bold, size, color) in enumerate(guia, 1):
        ws4.cell(row=r, column=1).value = text
        ws4.cell(row=r, column=1).font = Font(bold=bold, size=size, color=color or '000000')
    ws4.column_dimensions['A'].width = 85

    save(wb, f'{BASE}/dia09/datos_dia09_Access_BD.xlsx')

# ── Día 10: Funciones BD + IA ────────────────────────────────────────────────
def dia10():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Base_Ventas_BD'

    ws['A1'] = 'BASE DE DATOS VENTAS — FUNCIONES BD EN EXCEL'
    ws['A1'].font = Font(bold=True, size=12, color='34D399')
    ws.merge_cells('A1:H1')

    headers = ['VENDEDOR','CIUDAD','MES','PRODUCTO','CATEGORIA','CANTIDAD','PRECIO_UNIT','TOTAL_VENTA']
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i).value = h
    style_header(ws, 2, 8, '34D399', '001a06')

    vendedores2 = ['Carlos Martínez','Sandra Castro','Pedro Herrera','Diana Moreno']
    ciudades2 = ['Barranquilla','Bogotá','Cartagena','Barranquilla']
    meses_bd = ['Enero','Enero','Febrero','Febrero','Marzo','Marzo']
    prods2 = [('Papel','Papelería',18500),('Tóner','Tecnología',185000),
              ('USB','Tecnología',35000),('Carpeta','Archivos',8500)]

    random.seed(7)
    alts = ['E6FFF6','FFFFFF']
    for row_i in range(50):
        r = row_i + 3
        v_i = random.randint(0,3)
        p_i = random.randint(0,3)
        m_i = random.randint(0,5)
        cant = random.randint(1,15)
        prod, cat, precio = prods2[p_i]
        row_data = [vendedores2[v_i], ciudades2[v_i], meses_bd[m_i],
                    prod, cat, cant, precio, cant*precio]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in [7, 8]:
                cell.number_format = '"$"#,##0'
        style_data_row(ws, r, 8, alts[row_i % 2])

    set_col_widths(ws, [22, 14, 10, 14, 14, 10, 14, 16])
    freeze(ws)
    autofilter(ws, 'A2:H2')

    # Hoja de criterios y panel
    ws2 = wb.create_sheet('Panel_BD_Funciones')
    ws2['A1'] = 'PANEL DE ANÁLISIS — FUNCIONES BD'
    ws2['A1'].font = Font(bold=True, size=13, color='34D399')
    ws2.merge_cells('A1:D1')

    ws2['A3'] = 'RANGO DE CRITERIOS (editar celdas amarillas):'
    ws2['A3'].font = Font(bold=True)

    # Criterios encabezados (deben coincidir con los de la BD)
    criterio_headers = ['CIUDAD', 'MES', 'CATEGORIA']
    for i, h in enumerate(criterio_headers, 1):
        ws2.cell(row=4, column=i).value = h
        ws2.cell(row=4, column=i).font = Font(bold=True, color='34D399')
    ws2['A5'] = 'Barranquilla'
    ws2['B5'] = 'Enero'
    ws2['C5'] = ''
    for c in [1, 2, 3]:
        ws2.cell(row=5, column=c).fill = PatternFill('solid', fgColor='FFF2CC')

    ws2['A7'] = 'RESULTADOS:'
    ws2['A7'].font = Font(bold=True, size=11)

    resultados = [
        ('Total ventas con criterios:', '=BDSUMA(Base_Ventas_BD!A2:H52,"TOTAL_VENTA",A4:C5)', '"$"#,##0'),
        ('Nro de transacciones:', '=BDCONTAR(Base_Ventas_BD!A2:H52,"CANTIDAD",A4:C5)', '0'),
        ('Venta promedio:', '=BDPROMEDIO(Base_Ventas_BD!A2:H52,"TOTAL_VENTA",A4:C5)', '"$"#,##0'),
        ('Venta máxima:', '=BDMAX(Base_Ventas_BD!A2:H52,"TOTAL_VENTA",A4:C5)', '"$"#,##0'),
    ]
    for r, (label, formula, fmt) in enumerate(resultados, 8):
        ws2.cell(row=r, column=1).value = label
        ws2.cell(row=r, column=1).font = Font(bold=True, size=10)
        cell = ws2.cell(row=r, column=2, value=formula)
        cell.font = Font(bold=True, size=13, color='34D399')
        cell.number_format = fmt
        ws2.row_dimensions[r].height = 28

    ws2['A13'] = 'INSTRUCCIÓN: Cambia los valores en A5, B5, C5 para filtrar diferentes resultados.'
    ws2['A13'].font = Font(italic=True, color='888888')

    set_col_widths(ws2, [30, 24, 16, 16])

    ws3 = wb.create_sheet('GUIA_ACTIVIDAD')
    guia = [
        ('ACTIVIDAD PRÁCTICA — DÍA 10: Funciones BD + IA Administrativa', True, 13, '34D399'),
        ('', False, 10, None),
        ('PARTE 1 — FUNCIONES BD (30 min)', True, 11, '333333'),
        ('1. Ve a la hoja "Panel_BD_Funciones"', False, 10, None),
        ('2. Prueba cambiar la ciudad en A5 a "Bogotá" — ¿cambian todos los resultados?', False, 10, None),
        ('3. Cambia B5 a "Febrero" — ¿cuántas transacciones hubo en Bogotá en Febrero?', False, 10, None),
        ('4. Borra B5 — ahora muestra todas las ciudades sin filtro de mes', False, 10, None),
        ('5. Agrega un criterio O: pon "Cartagena" en A6 y repite los cálculos', False, 10, None),
        ('6. Añade BDMIN para encontrar la venta más pequeña', False, 10, None),
        ('', False, 10, None),
        ('PARTE 2 — TAREAS CON IA (35 min)', True, 11, '333333'),
        ('1. Abre ChatGPT o Copilot en el navegador', False, 10, None),
        ('2. Tarea A: pide que escriba un correo de cobro para el cliente "Distribuciones Caribe"', False, 10, None),
        ('   (factura #F-2025-042, $3.8M, 45 días vencida)', False, 10, None),
        ('3. Tarea B: pide una lista de 8 pasos para el proceso de cierre de ventas mensual', False, 10, None),
        ('4. Tarea C: copia una fórmula BD complicada y pide que la explique en español simple', False, 10, None),
        ('5. Guarda los 3 resultados en un documento Word — revisando y corrigiendo cada uno', False, 10, None),
        ('', False, 10, None),
        ('IMPORTANTE: Usa datos ficticios de ADMIN CARIBE, nunca datos reales de personas.', True, 10, 'C00000'),
        ('', False, 10, None),
        ('Guarda el Excel como: Actividad_D10_TuNombre.xlsx', False, 10, '7030A0'),
        ('Guarda el Word como: Tareas_IA_D10_TuNombre.docx', False, 10, '7030A0'),
    ]
    for r, (text, bold, size, color) in enumerate(guia, 1):
        ws3.cell(row=r, column=1).value = text
        ws3.cell(row=r, column=1).font = Font(bold=bold, size=size, color=color or '000000')
    ws3.column_dimensions['A'].width = 85

    save(wb, f'{BASE}/dia10/datos_dia10_FuncionesBD_IA.xlsx')

# ── Día 11: Proyecto Integrador ──────────────────────────────────────────────
def dia11():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DATOS_PROYECTO'

    ws['A1'] = 'BASE DE DATOS COMPLETA — PROYECTO INTEGRADOR ADMIN CARIBE S.A.S. 2025'
    ws['A1'].font = Font(bold=True, size=12, color='A78BFA')
    ws.merge_cells('A1:J1')

    headers = ['ID_VENTA','FECHA','VENDEDOR','CIUDAD','PRODUCTO','CATEGORIA',
               'CANTIDAD','PRECIO_UNIT','DESCUENTO%','TOTAL_NETO']
    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i).value = h
    style_header(ws, 2, 10, 'A78BFA', '0a0020')

    vendedores3 = ['Carlos Martínez','Sandra Castro','Pedro Herrera','Diana Moreno','Luis Vargas']
    ciudades3 = ['Barranquilla','Bogotá','Cartagena','Barranquilla','Medellín']
    prods3 = [
        ('Resma Papel','Papelería',18500),('Tóner HP','Tecnología',185000),
        ('USB 32GB','Tecnología',35000),('Carpeta AZ','Archivos',8500),
        ('Calculadora','Oficina',95000),('Grapadora','Herramientas',38000),
        ('Bolígrafo x12','Papelería',12000),('Folder Manila','Papelería',22000),
    ]
    meses_p = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
               'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']

    random.seed(99)
    alts = ['F3EEFF','FFFFFF']
    for row_i in range(120):
        r = row_i + 3
        v_i = random.randint(0, 4)
        p_i = random.randint(0, 7)
        mes = random.randint(1, 12)
        dia = random.randint(1, 28)
        cant = random.randint(1, 25)
        desc = random.choice([0, 0, 0, 5, 10, 15])
        prod, cat, precio = prods3[p_i]
        total = cant * precio * (1 - desc/100)
        row_data = [f'V{row_i+1:03d}', f'2025-{mes:02d}-{dia:02d}',
                    vendedores3[v_i], ciudades3[v_i], prod, cat,
                    cant, precio, desc/100, total]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in [8, 10]:
                cell.number_format = '"$"#,##0'
            elif c == 9:
                cell.number_format = '0%'
        style_data_row(ws, r, 10, alts[row_i % 2])

    set_col_widths(ws, [10, 13, 22, 14, 20, 14, 10, 14, 12, 16])
    freeze(ws)
    autofilter(ws, 'A2:J2')

    ws2 = wb.create_sheet('GUIA_PROYECTO')
    guia = [
        ('PROYECTO INTEGRADOR — DÍA 11: ENSAMBLAJE FINAL', True, 14, 'A78BFA'),
        ('', False, 10, None),
        ('Este archivo contiene 120 registros de ventas de todo el año 2025.', False, 10, None),
        ('Es la base de datos de tu PROYECTO INTEGRADOR.', False, 10, None),
        ('', False, 10, None),
        ('CHECKLIST DEL PROYECTO (a completar hoy):', True, 12, '333333'),
        ('', False, 10, None),
        ('□  EXCEL — Dashboard (25% de la nota)', True, 11, 'A78BFA'),
        ('   ✓ Mínimo 2 funciones avanzadas (BUSCARV + INDICE+COINCIDIR)', False, 10, None),
        ('   ✓ Tabla dinámica: ventas por ciudad y mes', False, 10, None),
        ('   ✓ Tabla dinámica: top vendedores con % del total', False, 10, None),
        ('   ✓ Gráfico de columnas + gráfico de líneas', False, 10, None),
        ('   ✓ Slicer de Ciudad conectado a ambas tablas', False, 10, None),
        ('   ✓ KPIs: Total ventas, Mejor vendedor, Venta promedio, % con descuento', False, 10, None),
        ('   ✓ Macro "ActualizarTodo" que refresca tablas dinámicas', False, 10, None),
        ('   ✓ Hoja DATOS protegida con contraseña', False, 10, None),
        ('   ✓ Guardado como .xlsm', False, 10, None),
        ('', False, 10, None),
        ('□  ACCESS — Base de datos (20% de la nota)', True, 11, 'A78BFA'),
        ('   ✓ 3 tablas: tbl_Clientes, tbl_Productos, tbl_Pedidos', False, 10, None),
        ('   ✓ Relaciones con integridad referencial', False, 10, None),
        ('   ✓ Consulta: pedidos pendientes de clientes de Barranquilla', False, 10, None),
        ('   ✓ Formulario de ingreso de clientes', False, 10, None),
        ('   ✓ Informe agrupado por ciudad', False, 10, None),
        ('', False, 10, None),
        ('□  POWERPOINT — Presentación ejecutiva (25% de la nota)', True, 11, 'A78BFA'),
        ('   ✓ 8-10 slides con patrón de diapositiva configurado', False, 10, None),
        ('   ✓ KPIs del trimestre destacados visualmente', False, 10, None),
        ('   ✓ 2 gráficos del dashboard de Excel incluidos', False, 10, None),
        ('   ✓ Notas del presentador en al menos 3 slides', False, 10, None),
        ('   ✓ Transiciones consistentes (solo Desvanecer o Empujar)', False, 10, None),
        ('', False, 10, None),
        ('□  DOCUMENTO IA (15% de la nota)', True, 11, 'A78BFA'),
        ('   ✓ Correo de cobro generado con IA y editado', False, 10, None),
        ('   ✓ Proceso de 8 pasos para cierre mensual (con IA)', False, 10, None),
        ('   ✓ Reflexión de 1 página: ¿qué aprendiste con la IA?', False, 10, None),
        ('', False, 10, None),
        ('□  EXPOSICIÓN ORAL (15% de la nota)', True, 11, 'A78BFA'),
        ('   ✓ Ensaya tu presentación de 5 minutos hoy', False, 10, None),
        ('   ✓ Mañana: llega 10 minutos antes de clase', False, 10, None),
        ('', False, 10, None),
        ('NOMBRES DE ARCHIVOS:', True, 11, '333333'),
        ('Dashboard_AdminCaribe_TuNombre.xlsm', False, 10, '7030A0'),
        ('AdminCaribe_BD_TuNombre.accdb', False, 10, '7030A0'),
        ('Presentacion_AdminCaribe_TuNombre.pptx', False, 10, '7030A0'),
        ('Tareas_IA_TuNombre.pdf', False, 10, '7030A0'),
    ]
    for r, (text, bold, size, color) in enumerate(guia, 1):
        ws2.cell(row=r, column=1).value = text
        ws2.cell(row=r, column=1).font = Font(bold=bold, size=size, color=color or '000000')
    ws2.column_dimensions['A'].width = 90

    save(wb, f'{BASE}/dia11/datos_dia11_Proyecto_Integrador.xlsx')

# ── Día 12: Evaluación ───────────────────────────────────────────────────────
def dia12():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'EVALUACION_DATOS'

    ws['A1'] = 'DATOS EVALUACIÓN PRÁCTICA — DÍA 12 — MÓDULO 11'
    ws['A1'].font = Font(bold=True, size=12, color='FBBF24')
    ws.merge_cells('A1:H1')
    ws['A2'] = 'Lee las instrucciones en la hoja "INSTRUCCIONES_EVALUACION" antes de empezar'
    ws['A2'].font = Font(italic=True, color='888888')
    ws.merge_cells('A2:H2')

    headers = ['ID','VENDEDOR','CIUDAD','MES','PRODUCTO','CATEGORIA','CANTIDAD','PRECIO']
    for i, h in enumerate(headers, 1):
        ws.cell(row=3, column=i).value = h
    style_header(ws, 3, 8, 'FBBF24', '1a0a00')

    vendedores_e = ['Ana López','Carlos Ruiz','María Pérez','José Torres']
    ciudades_e = ['Barranquilla','Bogotá','Barranquilla','Cartagena']
    prods_e = [('Papel A4','Papelería',18000),('Mouse Inalámbrico','Tecnología',65000),
               ('Silla Ergonómica','Mobiliario',450000),('Tóner','Tecnología',185000),
               ('Archivador','Archivos',125000)]
    meses_e = ['Enero','Enero','Febrero','Febrero','Marzo','Marzo']

    random.seed(2025)
    alts = ['FFF8DC','FFFFFF']
    for row_i in range(30):
        r = row_i + 4
        v_i = random.randint(0,3)
        p_i = random.randint(0,4)
        m_i = random.randint(0,5)
        cant = random.randint(1,10)
        prod, cat, precio = prods_e[p_i]
        row_data = [row_i+1, vendedores_e[v_i], ciudades_e[v_i],
                    meses_e[m_i], prod, cat, cant, precio]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 8:
                cell.number_format = '"$"#,##0'
        style_data_row(ws, r, 8, alts[row_i % 2])

    set_col_widths(ws, [6, 18, 15, 10, 22, 14, 10, 14])
    freeze(ws, 'A4')
    autofilter(ws, 'A3:H3')

    # Tabla de vendedores para BUSCARV
    ws_aux = wb.create_sheet('Tabla_Vendedores')
    ws_aux['A1'] = 'TABLA AUXILIAR — VENDEDORES'
    ws_aux['A1'].font = Font(bold=True, color='FBBF24')
    headers_aux = ['VENDEDOR','REGION','COMISION%','META_MENSUAL']
    for i, h in enumerate(headers_aux, 1):
        ws_aux.cell(row=2, column=i).value = h
    style_header(ws_aux, 2, 4, 'FBBF24', '1a0a00')

    for r_i, (v, region, pct, meta) in enumerate([
        ('Ana López','Caribe',4,15000000),
        ('Carlos Ruiz','Centro',3.5,20000000),
        ('María Pérez','Caribe',4,15000000),
        ('José Torres','Costa',3,12000000),
    ], 3):
        ws_aux.cell(row=r_i, column=1).value = v
        ws_aux.cell(row=r_i, column=2).value = region
        ws_aux.cell(row=r_i, column=3).value = pct/100
        ws_aux.cell(row=r_i, column=3).number_format = '0.0%'
        ws_aux.cell(row=r_i, column=4).value = meta
        ws_aux.cell(row=r_i, column=4).number_format = '"$"#,##0'
    set_col_widths(ws_aux, [20, 14, 14, 18])

    # Hoja respuestas estudiante
    ws_resp = wb.create_sheet('MIS_RESPUESTAS')
    ws_resp['A1'] = 'HOJA DE RESPUESTAS — EVALUACIÓN DÍA 12'
    ws_resp['A1'].font = Font(bold=True, size=13, color='FBBF24')
    ws_resp['A2'] = 'Nombre completo:'
    ws_resp['B2'].fill = PatternFill('solid', fgColor='FFF2CC')
    ws_resp['A3'] = 'Fecha:'
    ws_resp['B3'] = '2025-12-XX'
    ws_resp['B3'].fill = PatternFill('solid', fgColor='FFF2CC')

    tareas = [
        ('TAREA 1 — BUSCARV (10 min)', 'En la columna I de EVALUACION_DATOS, usa BUSCARV para traer la REGION del vendedor desde Tabla_Vendedores'),
        ('TAREA 2 — SUMAR.SI (10 min)', 'En esta hoja, calcula: total de ventas (cantidad×precio) de BARRANQUILLA en ENERO'),
        ('TAREA 3 — TABLA DINÁMICA (10 min)', 'Crea una tabla dinámica en hoja nueva: ventas totales por CIUDAD y MES'),
    ]
    for r_i, (tarea, desc) in enumerate(tareas, 5):
        ws_resp.cell(row=r_i*3-1, column=1).value = tarea
        ws_resp.cell(row=r_i*3-1, column=1).font = Font(bold=True, color='FBBF24', size=11)
        ws_resp.cell(row=r_i*3, column=1).value = desc
        ws_resp.cell(row=r_i*3, column=1).font = Font(italic=True, color='888888')
        ws_resp.cell(row=r_i*3+1, column=1).value = 'Tu respuesta/resultado:'
        ws_resp.cell(row=r_i*3+1, column=2).fill = PatternFill('solid', fgColor='FFF2CC')

    set_col_widths(ws_resp, [30, 50])

    ws_inst = wb.create_sheet('INSTRUCCIONES_EVALUACION')
    inst = [
        ('EVALUACIÓN PRÁCTICA — MÓDULO 11: HERRAMIENTAS OFIMÁTICAS II', True, 14, 'FBBF24'),
        ('', False, 10, None),
        ('INSTRUCCIONES GENERALES:', True, 12, '333333'),
        ('• Duración: 30 minutos', False, 10, None),
        ('• Trabajo individual — no se permite consultar a compañeros', False, 10, None),
        ('• SÍ puedes consultar tus apuntes y archivos del módulo', False, 10, None),
        ('• Guarda tu archivo cada 5 minutos (Ctrl+S)', False, 10, None),
        ('', False, 10, None),
        ('TAREA 1 — BUSCARV (10 puntos)', True, 11, 'FBBF24'),
        ('En la hoja "EVALUACION_DATOS", agrega la columna I con encabezado "REGION"', False, 10, None),
        ('Usa BUSCARV para traer la REGION de cada vendedor desde la hoja "Tabla_Vendedores"', False, 10, None),
        ('Si el vendedor no se encuentra, mostrar: "Sin región asignada"', False, 10, None),
        ('Pista: =IFERROR(BUSCARV(B4,Tabla_Vendedores!$A:$B,2,0),"Sin región asignada")', False, 10, '0070C0'),
        ('', False, 10, None),
        ('TAREA 2 — SUMAR.SI (10 puntos)', True, 11, 'FBBF24'),
        ('En la hoja "MIS_RESPUESTAS", calcula en la celda B11:', False, 10, None),
        ('   Total de ventas (CANTIDAD × PRECIO) de clientes de Barranquilla en Enero', False, 10, None),
        ('   Usa SUMAR.SI.CONJUNTO para filtrar por Ciudad=Barranquilla Y Mes=Enero', False, 10, None),
        ('   Recuerda: el total de cada venta es Cantidad × Precio (no hay columna total)', False, 10, None),
        ('', False, 10, None),
        ('TAREA 3 — TABLA DINÁMICA (10 puntos)', True, 11, 'FBBF24'),
        ('Crea una tabla dinámica en una hoja nueva llamada "TD_Evaluacion"', False, 10, None),
        ('   - FILAS: CIUDAD', False, 10, None),
        ('   - COLUMNAS: MES', False, 10, None),
        ('   - VALORES: suma de CANTIDAD (no de precio)', False, 10, None),
        ('   - Aplica un diseño de tabla que se vea ordenado', False, 10, None),
        ('', False, 10, None),
        ('ENTREGA:', True, 12, '333333'),
        ('Guarda como: Evaluacion_D12_TuNombre.xlsx', False, 10, '7030A0'),
        ('Entrega al instructor antes de que se acabe el tiempo.', False, 10, None),
    ]
    for r, (text, bold, size, color) in enumerate(inst, 1):
        ws_inst.cell(row=r, column=1).value = text
        ws_inst.cell(row=r, column=1).font = Font(bold=bold, size=size, color=color or '000000')
    ws_inst.column_dimensions['A'].width = 90

    save(wb, f'{BASE}/dia12/evaluacion_dia12_PRACTICA.xlsx')

# ── Run all ──────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print('Generando archivos Excel...')
    dia01()
    dia02()
    dia03()
    dia04()
    dia05()
    dia06()
    dia09()
    dia10()
    dia11()
    dia12()
    print('\n✓ Todos los archivos generados.')
